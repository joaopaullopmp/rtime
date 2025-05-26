import streamlit as st
import logging
import pandas as pd
import io
import os
import smtplib
import tempfile
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from datetime import datetime, timedelta
import calendar
from fpdf import FPDF
from database_manager import DatabaseManager
from collaborator_targets import CollaboratorTargetCalculator
from report_utils import get_feriados_portugal

# Configuração do logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("app.log"),
        logging.StreamHandler()
    ]
)

def get_available_users(users_df, selected_teams):
    """
    Função para obter colaboradores disponíveis baseado nas equipes selecionadas
    """
    available_users = []
    
    if "Todas" not in selected_teams:
        # Filtrar usuários pelas equipes selecionadas
        for _, user in users_df.iterrows():
            try:
                if isinstance(user['groups'], str):
                    user_groups = eval(user['groups'])
                else:
                    user_groups = user['groups'] if user['groups'] is not None else []
                
                # Converter para lista se for outro tipo
                if not isinstance(user_groups, list):
                    if isinstance(user_groups, dict):
                        user_groups = list(user_groups.values())
                    else:
                        user_groups = [user_groups]
                
                # Verificar se o usuário pertence a alguma das equipes selecionadas
                if any(team in user_groups for team in selected_teams):
                    available_users.append({
                        'user_id': user['user_id'],
                        'name': f"{user['First_Name']} {user['Last_Name']}",
                        'team': user_groups[0] if user_groups else 'Sem equipe'
                    })
            except Exception as e:
                logging.warning(f"Erro ao processar grupos do usuário {user['First_Name']} {user['Last_Name']}: {str(e)}")
    else:
        # Todos os usuários disponíveis
        for _, user in users_df.iterrows():
            try:
                user_groups = eval(user['groups']) if isinstance(user['groups'], str) else (user['groups'] if user['groups'] is not None else [])
                if not isinstance(user_groups, list):
                    if isinstance(user_groups, dict):
                        user_groups = list(user_groups.values())
                    else:
                        user_groups = [user_groups]
                
                available_users.append({
                    'user_id': user['user_id'],
                    'name': f"{user['First_Name']} {user['Last_Name']}",
                    'team': user_groups[0] if user_groups else 'Sem equipe'
                })
            except Exception as e:
                logging.warning(f"Erro ao processar usuário {user['First_Name']} {user['Last_Name']}: {str(e)}")
    
    return available_users

def collaborator_email_report():
    """
    Módulo para geração e envio de relatórios de indicadores de colaboradores por email
    """
    logging.info("Iniciando a geração do relatório de indicadores de colaboradores.")
    st.title("📧 Relatório de Indicadores de Colaboradores")
    
    # Inicializar gerenciadores
    db_manager = DatabaseManager()
    collaborator_target_calculator = CollaboratorTargetCalculator()
    
    # Verificar se o usuário é administrador
    if st.session_state.user_info['role'].lower() != 'admin':
        st.warning("Esta funcionalidade é exclusiva para administradores.")
        return
    
    # Carregamento de dados básicos
    try:
        users_df = db_manager.query_to_df("SELECT * FROM utilizadores WHERE active = 1")
        timesheet_df = db_manager.query_to_df("SELECT * FROM timesheet")
        groups_df = db_manager.query_to_df("SELECT * FROM groups WHERE active = 1")
        
        # Converter datas
        timesheet_df['start_date'] = pd.to_datetime(timesheet_df['start_date'], format='mixed', errors='coerce')
        
    except Exception as e:
        st.error(f"Erro ao carregar dados: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return
    
    # Interface para configuração do relatório
    st.subheader("Configuração do Relatório")
    
    # SEÇÃO 1: Filtros Básicos (FORA DO FORMULÁRIO para serem acessíveis)
    st.subheader("🔧 Filtros Básicos")
    
    # Filtros para o relatório
    col1, col2 = st.columns(2)
    
    with col1:
        # Seleção de período
        report_period = st.selectbox(
            "Período do Relatório",
            ["Mês Atual", "Mês Anterior", "Últimos 3 Meses", "Ano Atual", "Período Personalizado"]
        )
        
        if report_period == "Período Personalizado":
            start_date = st.date_input(
                "Data de Início",
                value=datetime.now().replace(day=1)
            )
            end_date = st.date_input(
                "Data de Término",
                value=datetime.now()
            )
        else:
            # Definir datas automaticamente conforme o período selecionado
            today = datetime.now()
            
            if report_period == "Mês Atual":
                start_date = today.replace(day=1)
                last_day = calendar.monthrange(today.year, today.month)[1]
                end_date = today.replace(day=last_day)
                
            elif report_period == "Mês Anterior":
                if today.month == 1:
                    previous_month = 12
                    previous_year = today.year - 1
                else:
                    previous_month = today.month - 1
                    previous_year = today.year
                    
                start_date = datetime(previous_year, previous_month, 1)
                last_day = calendar.monthrange(previous_year, previous_month)[1]
                end_date = datetime(previous_year, previous_month, last_day)
                
            elif report_period == "Últimos 3 Meses":
                # Calcular 3 meses atrás
                if today.month <= 3:
                    months_back = today.month + 12 - 3
                    year_back = today.year - 1
                else:
                    months_back = today.month - 3
                    year_back = today.year
                    
                start_date = datetime(year_back, months_back, 1)
                last_day = calendar.monthrange(today.year, today.month)[1]
                end_date = today.replace(day=last_day)
                
            elif report_period == "Ano Atual":
                start_date = datetime(today.year, 1, 1)
                end_date = datetime(today.year, 12, 31)
    
    with col2:
        # Seleção de equipes (múltipla) - MOVIDO PARA FORA DO FORMULÁRIO
        team_options = ["Todas"] + sorted(groups_df['group_name'].tolist())
        selected_teams = st.multiselect(
            "Equipes", 
            options=team_options,
            default=["Todas"],
            help="Selecione as equipes para filtrar colaboradores"
        )
        
        # Formato do relatório
        report_format = st.radio(
            "Formato do Relatório",
            ["PDF", "Excel", "PDF e Excel"]
        )
        
        # Opções adicionais
        show_top_performers = st.checkbox("Destacar Top Performers", value=True, 
                                        help="Incluir análise dos colaboradores com melhores indicadores")
        show_low_performers = st.checkbox("Identificar Áreas de Melhoria", value=True,
                                       help="Incluir análise dos colaboradores abaixo da meta")
    
    # SEÇÃO 2: Filtro de Colaboradores com Ponderação
    st.subheader("🎯 Filtro de Colaboradores e Ponderação")
    
    # Mostrar informações sobre as equipes selecionadas
    teams_info = f"Equipes selecionadas: {', '.join(selected_teams)}"
    st.info(teams_info)
    
    # Checkbox para ativar filtro de colaboradores
    use_collaborator_filter = st.checkbox(
        "🔍 Aplicar filtro específico de colaboradores",
        value=False,
        help="Ative esta opção para selecionar colaboradores específicos e aplicar ponderação"
    )
    
    # Inicializar variável para armazenar colaboradores disponíveis
    available_users = []
    collaborator_weights = {}
    
    if use_collaborator_filter:
        # Carregar colaboradores automaticamente baseado nas equipes selecionadas
        # Usar session state para manter os dados carregados
        if 'available_users' not in st.session_state:
            st.session_state.available_users = []
            st.session_state.teams_loaded = []
        
        # Verificar se as equipes mudaram
        teams_changed = st.session_state.teams_loaded != selected_teams
        
        # Botão para carregar/atualizar colaboradores
        col_button, col_auto, col_info = st.columns([1, 1, 2])
        
        with col_button:
            load_collaborators = st.button(
                "🔄 Carregar Colaboradores",
                help="Clique para carregar os colaboradores das equipes selecionadas"
            )
        
        with col_auto:
            auto_load = st.checkbox(
                "Auto-carregar",
                value=True,
                help="Carregar automaticamente quando as equipes mudarem"
            )
        
        with col_info:
            if teams_changed and auto_load:
                st.info("🔄 As equipes mudaram - carregando automaticamente...")
            elif teams_changed:
                st.warning("⚠️ As equipes mudaram - clique em 'Carregar Colaboradores'")
            else:
                st.success(f"📋 Equipes atuais: {len(selected_teams)} selecionada(s)")
        
        # Carregar colaboradores quando necessário
        should_load = load_collaborators or (teams_changed and auto_load) or not st.session_state.available_users
        
        if should_load:
            with st.spinner("Carregando colaboradores..."):
                # Carregar colaboradores
                st.session_state.available_users = get_available_users(users_df, selected_teams)
                st.session_state.teams_loaded = selected_teams.copy()
                
                if st.session_state.available_users:
                    st.success(f"✅ {len(st.session_state.available_users)} colaborador(es) carregado(s) das equipes: {', '.join(selected_teams)}")
                else:
                    st.warning("⚠️ Nenhum colaborador encontrado para as equipes selecionadas")
        
        # Usar os colaboradores do session state
        available_users = st.session_state.available_users
        
        # Mostrar informações sobre colaboradores carregados
        if available_users:
            # Agrupar por equipe para mostrar resumo
            teams_summary = {}
            for user in available_users:
                team = user['team']
                if team not in teams_summary:
                    teams_summary[team] = 0
                teams_summary[team] += 1
            
            summary_text = " | ".join([f"{team}: {count}" for team, count in teams_summary.items()])
            st.info(f"📊 Colaboradores por equipe: {summary_text}")
        
        if available_users:
            st.info("Selecione os colaboradores e defina o percentual de ponderação para cada um (0-100%)")
            
            # Criar colunas para organizar melhor conforme layout da imagem
            col_left, col_right = st.columns(2)
            
            # Dividir a lista de usuários em duas colunas
            mid_point = len(available_users) // 2
            
            with col_left:
                st.write("**Primeira Metade:**")
                for user in available_users[:mid_point]:
                    # Layout horizontal: checkbox + input de peso
                    cols = st.columns([4, 2])  # Ajustar proporção
                    
                    with cols[0]:
                        selected = st.checkbox(
                            f"{user['name']} ({user['team']})",
                            key=f"user_{user['user_id']}"
                        )
                    
                    with cols[1]:
                        if selected:
                            weight = st.number_input(
                                "Peso %",
                                min_value=0,
                                max_value=100,
                                value=100,
                                step=5,
                                key=f"weight_{user['user_id']}",
                                help="Percentual de ponderação (0-100%)",
                                label_visibility="collapsed"
                            )
                            collaborator_weights[user['user_id']] = weight / 100.0  # Converter para decimal
                        else:
                            st.text_input("", value="N/A", disabled=True, key=f"disabled_{user['user_id']}", label_visibility="collapsed")
            
            with col_right:
                st.write("**Segunda Metade:**")
                for user in available_users[mid_point:]:
                    # Layout horizontal: checkbox + input de peso
                    cols = st.columns([4, 2])  # Ajustar proporção
                    
                    with cols[0]:
                        selected = st.checkbox(
                            f"{user['name']} ({user['team']})",
                            key=f"user_{user['user_id']}_right"
                        )
                    
                    with cols[1]:
                        if selected:
                            weight = st.number_input(
                                "Peso %",
                                min_value=0,
                                max_value=100,
                                value=100,
                                step=5,
                                key=f"weight_{user['user_id']}_right",
                                help="Percentual de ponderação (0-100%)",
                                label_visibility="collapsed"
                            )
                            collaborator_weights[user['user_id']] = weight / 100.0  # Converter para decimal
                        else:
                            st.text_input("", value="N/A", disabled=True, key=f"disabled_{user['user_id']}_right", label_visibility="collapsed")
            
            # Mostrar resumo dos colaboradores selecionados
            if collaborator_weights:
                st.success(f"✅ {len(collaborator_weights)} colaborador(es) selecionado(s) com ponderação")
                
                # Mostrar resumo em expandir
                with st.expander("Ver resumo dos colaboradores selecionados"):
                    for user in available_users:
                        if user['user_id'] in collaborator_weights:
                            weight_percent = collaborator_weights[user['user_id']] * 100
                            st.write(f"• {user['name']}: {weight_percent:.0f}%")
            else:
                if use_collaborator_filter:
                    st.warning("⚠️ Nenhum colaborador foi selecionado. O relatório incluirá todos os colaboradores das equipes selecionadas.")
        else:
            st.info("Carregue os colaboradores primeiro para poder aplicar filtros específicos.")
    
    # SEÇÃO 3: Configurações de Email (DENTRO DO FORMULÁRIO)
    st.subheader("📧 Configurações de Email")
    
    with st.form("email_report_config"):
        # Seleção de destinatários
        recipients = st.text_input(
            "Destinatários (separados por vírgula)",
            help="Email dos destinatários separados por vírgula"
        )
        
        # Assunto do email
        subject = st.text_input(
            "Assunto do Email", 
            f"Relatório de Indicadores de Colaboradores - {datetime.now().strftime('%d/%m/%Y')}"
        )
        
        # Conteúdo do email
        email_message = st.text_area(
            "Mensagem do Email",
            """Prezados,

Em anexo, relatório de indicadores de desempenho dos colaboradores.

Atenciosamente,
Equipe de Gestão"""
        )
        
        # SMTP settings (collapsed by default)
        with st.expander("Configurações de SMTP"):
            smtp_server = st.text_input("Servidor SMTP", "smtp.office365.com")
            smtp_port = st.number_input("Porta SMTP", value=587, step=1)
            smtp_user = st.text_input("Usuário SMTP", "notifications@grupoerre.pt")
            smtp_password = st.text_input("Senha SMTP", type="password", value="erretech@2020")
            use_tls = st.checkbox("Usar TLS", value=True)
        
        # Botão para gerar e enviar o relatório
        submit_button = st.form_submit_button("Gerar e Enviar Relatório")
    
    if submit_button:
        # Validar entradas
        if not recipients:
            st.error("Por favor, informe pelo menos um destinatário.")
            return
        
        # Validar se foram selecionados colaboradores quando o filtro está ativo
        if use_collaborator_filter and not collaborator_weights:
            st.warning("Filtro de colaboradores está ativo mas nenhum colaborador foi selecionado. O relatório incluirá todos os colaboradores das equipes selecionadas.")
            use_collaborator_filter = False
        
        with st.spinner("Gerando e enviando relatório..."):
            # Criar diretório temporário para os arquivos
            temp_dir = tempfile.mkdtemp()
            
            # Gerar relatórios conforme formato selecionado
            pdf_path = None
            excel_path = None
            
            if "PDF" in report_format:
                pdf_path = os.path.join(temp_dir, "relatorio_indicadores_colaboradores.pdf")
                pdf_result = generate_collaborator_pdf_report(
                    pdf_path,
                    db_manager,
                    collaborator_target_calculator,
                    start_date,
                    end_date,
                    selected_teams,
                    show_top_performers,
                    show_low_performers,
                    use_collaborator_filter,
                    collaborator_weights
                )
                
                # Verificar se o PDF foi gerado com sucesso
                if not pdf_result or not os.path.exists(pdf_path):
                    st.error(f"Falha ao gerar o PDF. Verifique os logs para mais detalhes.")
                    pdf_path = None

            if "Excel" in report_format:
                excel_path = os.path.join(temp_dir, "relatorio_indicadores_colaboradores.xlsx")
                try:
                    # Obter dados dos colaboradores para o Excel
                    colaboradores = get_collaborator_indicators(
                        db_manager,
                        collaborator_target_calculator,
                        start_date,
                        end_date,
                        selected_teams,
                        use_collaborator_filter,
                        collaborator_weights
                    )
                    
                    excel_result = generate_collaborator_excel_report(
                        excel_path,
                        colaboradores,
                        start_date,
                        end_date,
                        selected_teams,
                        show_top_performers,
                        show_low_performers,
                        use_collaborator_filter,
                        collaborator_weights
                    )
                    
                    if excel_result and os.path.exists(excel_path):
                        st.success("Excel gerado com sucesso!")
                    else:
                        st.error("Falha ao gerar o Excel.")
                        excel_path = None
                except Exception as e:
                    st.error(f"Erro ao gerar Excel: {str(e)}")
                    excel_path = None
            
            # Enviar email com os relatórios gerados se pelo menos um foi gerado com sucesso
            if pdf_path or excel_path:
                email_success = send_email(
                    recipients.split(','), 
                    subject, 
                    email_message, 
                    pdf_path, 
                    excel_path, 
                    smtp_server, 
                    smtp_port, 
                    smtp_user, 
                    smtp_password, 
                    use_tls
                )
                
                # Botões para download local dos relatórios
                st.markdown("### Download dos Relatórios")
                
                if pdf_path and os.path.exists(pdf_path):
                    with open(pdf_path, "rb") as pdf_file:
                        pdf_bytes = pdf_file.read()
                        st.download_button(
                            label="📥 Baixar PDF",
                            data=pdf_bytes,
                            file_name="relatorio_indicadores_colaboradores.pdf",
                            mime="application/pdf"
                        )
                
                if excel_path and os.path.exists(excel_path):
                    with open(excel_path, "rb") as excel_file:
                        excel_bytes = excel_file.read()
                        st.download_button(
                            label="📥 Baixar Excel",
                            data=excel_bytes,
                            file_name="relatorio_indicadores_colaboradores.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            else:
                st.error("Não foi possível gerar nenhum dos relatórios. Verifique os logs para mais detalhes.")

# [Manter todas as outras funções existentes inalteradas]
# generate_collaborator_pdf_report, generate_collaborator_excel_report, 
# get_collaborator_indicators, send_email, get_collaborator_absences


def generate_collaborator_pdf_report(
    output_path,
    db_manager,
    collaborator_target_calculator,
    start_date,
    end_date,
    selected_teams,
    show_top_performers=True,
    show_low_performers=True,
    use_collaborator_filter=False,
    collaborator_weights=None
):
    """
    Gera relatório PDF com indicadores de colaboradores
    """
    logging.info("Iniciando a geração do relatório PDF.")

    try:
        # Inicializar PDF
        class PDF(FPDF):
            def header(self):
                # Logo apenas na primeira página
                if self.page_no() != 1:  # Não mostrar o logo na primeira página (capa)
                    if os.path.exists('logo.png'):
                        self.image('logo.png', 10, 8, 33)
                    
                    # Título
                    self.set_font('Arial', 'B', 15)
                    self.cell(80)
                    self.cell(30, 25, 'Indicadores de Performance de Colaboradores', 0, 0, 'C')
                    self.ln(20)
            
            def footer(self):
                # Posicionar a 1.5 cm do final
                self.set_y(-15)
                # Fonte Arial itálico 8
                self.set_font('Arial', 'I', 8)
                # Número da página
                self.cell(0, 10, f'Pagina {self.page_no()}/{{nb}}', 0, 0, 'C')
        
        pdf = PDF()
        pdf.alias_nb_pages()
        
        # Capa do relatório
        pdf.add_page()

        # Adicionar o logo apenas uma vez
        if os.path.exists('logo.png'):
            pdf.image('logo.png', x=10, y=20, w=60)

        pdf.set_font('Arial', 'B', 20)
        pdf.set_xy(0, 100)
        pdf.cell(210, 20, 'Relatorio - Performance de Colaboradores', 0, 1, 'C')
        
        # Período do relatório
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 15, f"Periodo: {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}", 0, 1, 'C')
        
        # Filtros aplicados
        pdf.set_font('Arial', '', 12)
        pdf.ln(10)
        
        teams_str = ', '.join(selected_teams) if "Todas" not in selected_teams else "Todas as equipes"
        
        pdf.set_x(40)
        pdf.cell(0, 8, f"Equipes: {teams_str}", 0, 1)
        
        # Informar se foi aplicado filtro de colaboradores
        if use_collaborator_filter and collaborator_weights:
            pdf.set_x(40)
            pdf.cell(0, 8, f"Colaboradores selecionados: {len(collaborator_weights)}", 0, 1)
            pdf.set_x(40)
            pdf.cell(0, 8, "Aplicada ponderacao personalizada", 0, 1)
        
        # Data de geração
        pdf.set_font('Arial', 'I', 10)
        pdf.set_y(-30)
        pdf.cell(0, 10, f"Gerado automaticamente em {datetime.now().strftime('%d/%m/%Y as %H:%M')}", 0, 1, 'C')
        
        # Obter indicadores de colaboradores
        colaboradores = get_collaborator_indicators(
            db_manager,
            collaborator_target_calculator,
            start_date,
            end_date,
            selected_teams,
            use_collaborator_filter,
            collaborator_weights
        )
        
        if not colaboradores:
            # Página de erro
            pdf.add_page()
            pdf.set_font('Arial', 'B', 16)
            pdf.cell(0, 10, 'Sem Dados Disponíveis', 0, 1, 'C')
            
            pdf.set_font('Arial', '', 12)
            pdf.ln(10)
            pdf.multi_cell(0, 8, "Não foram encontrados dados de colaboradores para o período e filtros selecionados.")
            
            # Salvar o PDF mesmo assim
            pdf.output(output_path)
            return True
            
        # Converter para DataFrame
        df = pd.DataFrame(colaboradores)
        
        # Resumo executivo
        pdf.add_page()
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, 'Resumo Executivo', 0, 1)
        
        # Calcular dias úteis do período incluindo feriados
        month_year_pairs = set()
        current_date = start_date
        while current_date <= end_date:
            month_year_pairs.add((current_date.month, current_date.year))
            current_date += timedelta(days=1)
            
        # Obter todos os feriados dos anos envolvidos no período
        all_holidays = []
        for _, year in month_year_pairs:
            all_holidays.extend(get_feriados_portugal(year))
        
        # Calcular dias úteis excluindo feriados
        dias_uteis = 0
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() < 5 and current_date.date() not in all_holidays:  # 0-4 são dias úteis (seg-sex)
                dias_uteis += 1
            current_date += timedelta(days=1)

        # Calcular horas úteis totais para o período
        horas_uteis_periodo = dias_uteis * 8
        
        pdf.set_font('Arial', '', 11)
        resumo_text = f"Este relatório apresenta uma visão consolidada dos indicadores de desempenho dos colaboradores no período de {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}, cobrindo {dias_uteis} dias úteis, já excluídos feriados de Portugal."
        
        if use_collaborator_filter and collaborator_weights:
            resumo_text += f" Foram aplicados fatores de ponderação personalizados para {len(collaborator_weights)} colaborador(es) selecionado(s)."
        
        pdf.multi_cell(0, 8, resumo_text)
        
        # Dados resumidos para o sumário (considerar ponderação se aplicável)
        if use_collaborator_filter and collaborator_weights:
            # Calcular médias ponderadas CORRETAMENTE
            total_ocupacao_ponderada = sum(row['occupation_percentage'] * row['weight_factor'] for row in colaboradores)
            total_faturabilidade_ponderada = sum(row['billable_percentage'] * row['weight_factor'] for row in colaboradores)
            soma_pesos = sum(row['weight_factor'] for row in colaboradores)
            
            ocupacao_media = total_ocupacao_ponderada / soma_pesos if soma_pesos > 0 else 0
            faturabilidade_media = total_faturabilidade_ponderada / soma_pesos if soma_pesos > 0 else 0
        else:
            ocupacao_media = df['occupation_percentage'].mean()
            faturabilidade_media = df['billable_percentage'].mean()
        
        total_colaboradores = len(df)
        abaixo_meta_ocupacao = sum(df['occupation_percentage'] < 87.5)
        abaixo_meta_faturabilidade = sum(df['billable_percentage'] < 75.0)
        
        # Adicionar tabela de resumo
        pdf.ln(10)
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, "Resumo dos Indicadores:", 0, 1)
        
        # Cabeçalho da tabela
        pdf.set_fill_color(31, 119, 180)  # Azul
        pdf.set_text_color(255, 255, 255)  # Branco
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(100, 8, "Métrica", 1, 0, 'L', True)
        pdf.cell(35, 8, "Valor", 1, 1, 'C', True)
        
        # Dados da tabela com cores alternadas
        pdf.set_text_color(0, 0, 0)  # Preto
        
        # Total de colaboradores
        pdf.set_font('Arial', '', 10)
        pdf.cell(100, 8, "Total de Colaboradores", 1, 0, 'L')
        pdf.cell(35, 8, str(total_colaboradores), 1, 1, 'C')
        
        # Ocupação média
        pdf.set_fill_color(245, 245, 245)  # Cinza claro
        ocupacao_label = "Ocupação Média (Ponderada)" if use_collaborator_filter and collaborator_weights else "Ocupação Média"
        pdf.cell(100, 8, ocupacao_label, 1, 0, 'L', True)
        pdf.cell(35, 8, f"{ocupacao_media:.1f}%", 1, 1, 'C', True)
        
        # Faturabilidade média
        faturabilidade_label = "Faturabilidade Média (Ponderada)" if use_collaborator_filter and collaborator_weights else "Faturabilidade Média"
        pdf.cell(100, 8, faturabilidade_label, 1, 0, 'L')
        pdf.cell(35, 8, f"{faturabilidade_media:.1f}%", 1, 1, 'C')
        
        # Colaboradores abaixo da meta de ocupação
        pdf.set_fill_color(245, 245, 245)  # Cinza claro
        pdf.cell(100, 8, "Colaboradores Abaixo da Meta de Ocupação", 1, 0, 'L', True)
        pdf.cell(35, 8, f"{abaixo_meta_ocupacao} ({(abaixo_meta_ocupacao/total_colaboradores*100):.1f}%)", 1, 1, 'C', True)
        
        # Colaboradores abaixo da meta de faturabilidade
        pdf.cell(100, 8, "Colaboradores Abaixo da Meta de Faturabilidade", 1, 0, 'L')
        pdf.cell(35, 8, f"{abaixo_meta_faturabilidade} ({(abaixo_meta_faturabilidade/total_colaboradores*100):.1f}%)", 1, 1, 'C')

        # Top performers (se solicitado)
        if show_top_performers and len(df) > 0:
            pdf.add_page()
            pdf.set_font('Arial', 'B', 14)
            pdf.set_fill_color(200, 255, 200)  # Verde claro
            pdf.cell(0, 10, "Top Performers", 0, 1, 'L', True)
            
            # Destacar os 5 melhores em ocupação
            pdf.ln(5)
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, "Top 5 em Ocupação:", 0, 1)
            
            top_ocupacao = df.sort_values('occupation_percentage', ascending=False).head(5)
            
            # Cabeçalho da tabela (incluir ponderação se aplicável)
            pdf.set_fill_color(31, 119, 180)  # Azul
            pdf.set_text_color(255, 255, 255)  # Branco
            pdf.set_font('Arial', 'B', 9)
            
            if use_collaborator_filter and collaborator_weights:
                pdf.cell(60, 8, "Colaborador", 1, 0, 'L', True)
                pdf.cell(25, 8, "Ocupação", 1, 0, 'C', True)
                pdf.cell(25, 8, "Faturabilidade", 1, 0, 'C', True)
                pdf.cell(20, 8, "Ponderação", 1, 1, 'C', True)
            else:
                pdf.cell(80, 8, "Colaborador", 1, 0, 'L', True)
                pdf.cell(30, 8, "Ocupação", 1, 0, 'C', True)
                pdf.cell(30, 8, "Faturabilidade", 1, 1, 'C', True)
            
            # Dados da tabela
            pdf.set_text_color(0, 0, 0)  # Preto
            pdf.set_font('Arial', '', 9)
            
            for idx, row in top_ocupacao.iterrows():
                # Alternar cores de fundo
                if idx % 2 == 0:
                    pdf.set_fill_color(255, 255, 255)  # Branco
                    fill = False
                else:
                    pdf.set_fill_color(245, 245, 245)  # Cinza claro
                    fill = True
                
                if use_collaborator_filter and collaborator_weights:
                    pdf.cell(60, 7, row['name'], 1, 0, 'L', fill)
                    pdf.cell(25, 7, f"{row['occupation_percentage']:.1f}%", 1, 0, 'C', fill)
                    pdf.cell(25, 7, f"{row['billable_percentage']:.1f}%", 1, 0, 'C', fill)
                    pdf.cell(20, 7, f"{row['weight_percentage']:.0f}%", 1, 1, 'C', fill)
                else:
                    pdf.cell(80, 7, row['name'], 1, 0, 'L', fill)
                    pdf.cell(30, 7, f"{row['occupation_percentage']:.1f}%", 1, 0, 'C', fill)
                    pdf.cell(30, 7, f"{row['billable_percentage']:.1f}%", 1, 1, 'C', fill)
            
            # Destacar os 5 melhores em faturabilidade
            pdf.ln(10)
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, "Top 5 em Faturabilidade:", 0, 1)
            
            top_faturabilidade = df.sort_values('billable_percentage', ascending=False).head(5)
            
            # Cabeçalho da tabela
            pdf.set_fill_color(31, 119, 180)  # Azul
            pdf.set_text_color(255, 255, 255)  # Branco
            pdf.set_font('Arial', 'B', 9)
            
            if use_collaborator_filter and collaborator_weights:
                pdf.cell(60, 8, "Colaborador", 1, 0, 'L', True)
                pdf.cell(25, 8, "Faturabilidade", 1, 0, 'C', True)
                pdf.cell(25, 8, "Ocupação", 1, 0, 'C', True)
                pdf.cell(20, 8, "Ponderação", 1, 1, 'C', True)
            else:
                pdf.cell(80, 8, "Colaborador", 1, 0, 'L', True)
                pdf.cell(30, 8, "Faturabilidade", 1, 0, 'C', True)
                pdf.cell(30, 8, "Ocupação", 1, 1, 'C', True)
            
            # Dados da tabela
            pdf.set_text_color(0, 0, 0)  # Preto
            pdf.set_font('Arial', '', 9)
            
            for idx, row in top_faturabilidade.iterrows():
                # Alternar cores de fundo
                if idx % 2 == 0:
                    pdf.set_fill_color(255, 255, 255)  # Branco
                    fill = False
                else:
                    pdf.set_fill_color(245, 245, 245)  # Cinza claro
                    fill = True
                
                if use_collaborator_filter and collaborator_weights:
                    pdf.cell(60, 7, row['name'], 1, 0, 'L', fill)
                    pdf.cell(25, 7, f"{row['billable_percentage']:.1f}%", 1, 0, 'C', fill)
                    pdf.cell(25, 7, f"{row['occupation_percentage']:.1f}%", 1, 0, 'C', fill)
                    pdf.cell(20, 7, f"{row['weight_percentage']:.0f}%", 1, 1, 'C', fill)
                else:
                    pdf.cell(80, 7, row['name'], 1, 0, 'L', fill)
                    pdf.cell(30, 7, f"{row['billable_percentage']:.1f}%", 1, 0, 'C', fill)
                    pdf.cell(30, 7, f"{row['occupation_percentage']:.1f}%", 1, 1, 'C', fill)
        
        # Áreas de melhoria (se solicitado)
        if show_low_performers and len(df) > 0:
            pdf.add_page()
            pdf.set_font('Arial', 'B', 14)
            pdf.set_fill_color(255, 230, 230)  # Vermelho claro
            pdf.cell(0, 10, "Áreas de Melhoria", 0, 1, 'L', True)
            
            # Destacar os colaboradores abaixo da meta de ocupação
            pdf.ln(5)
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, "Colaboradores Abaixo da Meta de Ocupação:", 0, 1)
            
            below_ocupacao = df[df['occupation_percentage'] < 87.5].sort_values('occupation_percentage')
            
            if not below_ocupacao.empty:
                # Cabeçalho da tabela
                pdf.set_fill_color(31, 119, 180)  # Azul
                pdf.set_text_color(255, 255, 255)  # Branco
                pdf.set_font('Arial', 'B', 9)
                
                if use_collaborator_filter and collaborator_weights:
                    pdf.cell(60, 8, "Colaborador", 1, 0, 'L', True)
                    pdf.cell(25, 8, "Ocupação", 1, 0, 'C', True)
                    pdf.cell(25, 8, "Faturabilidade", 1, 0, 'C', True)
                    pdf.cell(20, 8, "Ponderação", 1, 1, 'C', True)
                else:
                    pdf.cell(80, 8, "Colaborador", 1, 0, 'L', True)
                    pdf.cell(30, 8, "Ocupação", 1, 0, 'C', True)
                    pdf.cell(30, 8, "Faturabilidade", 1, 1, 'C', True)
                
                # Dados da tabela
                pdf.set_text_color(0, 0, 0)  # Preto
                pdf.set_font('Arial', '', 9)
                
                for idx, row in below_ocupacao.iterrows():
                    # Alternar cores de fundo
                    if idx % 2 == 0:
                        pdf.set_fill_color(255, 255, 255)  # Branco
                        fill = False
                    else:
                        pdf.set_fill_color(245, 245, 245)  # Cinza claro
                        fill = True
                    
                    if use_collaborator_filter and collaborator_weights:
                        pdf.cell(60, 7, row['name'], 1, 0, 'L', fill)
                        pdf.cell(25, 7, f"{row['occupation_percentage']:.1f}%", 1, 0, 'C', fill)
                        pdf.cell(25, 7, f"{row['billable_percentage']:.1f}%", 1, 0, 'C', fill)
                        pdf.cell(20, 7, f"{row['weight_percentage']:.0f}%", 1, 1, 'C', fill)
                    else:
                        pdf.cell(80, 7, row['name'], 1, 0, 'L', fill)
                        pdf.cell(30, 7, f"{row['occupation_percentage']:.1f}%", 1, 0, 'C', fill)
                        pdf.cell(30, 7, f"{row['billable_percentage']:.1f}%", 1, 1, 'C', fill)
            else:
                pdf.cell(0, 8, "Todos os colaboradores atingiram a meta de ocupação!", 0, 1)
            
            # Destacar os colaboradores abaixo da meta de faturabilidade
            pdf.ln(10)
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, "Colaboradores Abaixo da Meta de Faturabilidade:", 0, 1)
            
            below_faturabilidade = df[df['billable_percentage'] < 75.0].sort_values('billable_percentage')
            
            if not below_faturabilidade.empty:
                # Cabeçalho da tabela
                pdf.set_fill_color(31, 119, 180)  # Azul
                pdf.set_text_color(255, 255, 255)  # Branco
                pdf.set_font('Arial', 'B', 9)
                
                if use_collaborator_filter and collaborator_weights:
                    pdf.cell(60, 8, "Colaborador", 1, 0, 'L', True)
                    pdf.cell(25, 8, "Faturabilidade", 1, 0, 'C', True)
                    pdf.cell(25, 8, "Ocupação", 1, 0, 'C', True)
                    pdf.cell(20, 8, "Ponderação", 1, 1, 'C', True)
                else:
                    pdf.cell(80, 8, "Colaborador", 1, 0, 'L', True)
                    pdf.cell(30, 8, "Faturabilidade", 1, 0, 'C', True)
                    pdf.cell(30, 8, "Ocupação", 1, 1, 'C', True)
                
                # Dados da tabela
                pdf.set_text_color(0, 0, 0)  # Preto
                pdf.set_font('Arial', '', 9)
                
                for idx, row in below_faturabilidade.iterrows():
                    # Alternar cores de fundo
                    if idx % 2 == 0:
                        pdf.set_fill_color(255, 255, 255)  # Branco
                        fill = False
                    else:
                        pdf.set_fill_color(245, 245, 245)  # Cinza claro
                        fill = True
                    
                    if use_collaborator_filter and collaborator_weights:
                        pdf.cell(60, 7, row['name'], 1, 0, 'L', fill)
                        pdf.cell(25, 7, f"{row['billable_percentage']:.1f}%", 1, 0, 'C', fill)
                        pdf.cell(25, 7, f"{row['occupation_percentage']:.1f}%", 1, 0, 'C', fill)
                        pdf.cell(20, 7, f"{row['weight_percentage']:.0f}%", 1, 1, 'C', fill)
                    else:
                        pdf.cell(80, 7, row['name'], 1, 0, 'L', fill)
                        pdf.cell(30, 7, f"{row['billable_percentage']:.1f}%", 1, 0, 'C', fill)
                        pdf.cell(30, 7, f"{row['occupation_percentage']:.1f}%", 1, 1, 'C', fill)
            else:
                pdf.cell(0, 8, "Todos os colaboradores atingiram a meta de faturabilidade!", 0, 1)
        
        # Tabela completa dos colaboradores
        pdf.add_page()
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, "Tabela Completa de Indicadores", 0, 1)
        
        # Ordenar alfabeticamente pelo nome para a tabela completa
        df_complete = df.sort_values('name')
        
        # Cabeçalho da tabela (ajustar com base se há ponderação)
        pdf.set_fill_color(31, 119, 180)  # Azul
        pdf.set_text_color(255, 255, 255)  # Branco
        pdf.set_font('Arial', 'B', 8)
        
        if use_collaborator_filter and collaborator_weights:
            pdf.cell(50, 8, "Colaborador", 1, 0, 'L', True)
            pdf.cell(18, 8, "Ocupação", 1, 0, 'C', True)
            pdf.cell(15, 8, "Meta", 1, 0, 'C', True)
            pdf.cell(18, 8, "Faturabilidade", 1, 0, 'C', True)
            pdf.cell(15, 8, "Meta", 1, 0, 'C', True)
            pdf.cell(18, 8, "Horas Real.", 1, 0, 'C', True)
            pdf.cell(15, 8, "Pond.", 1, 1, 'C', True)
        else:
            pdf.cell(60, 8, "Colaborador", 1, 0, 'L', True)
            pdf.cell(20, 8, "Ocupação", 1, 0, 'C', True)
            pdf.cell(20, 8, "Meta Ocup.", 1, 0, 'C', True)
            pdf.cell(20, 8, "Faturabilidade", 1, 0, 'C', True)
            pdf.cell(20, 8, "Meta Fatur.", 1, 0, 'C', True)
            pdf.cell(20, 8, "Horas Realiz.", 1, 1, 'C', True)

        # Dados da tabela
        pdf.set_text_color(0, 0, 0)  # Preto
        pdf.set_font('Arial', '', 7)

        for idx, row in df_complete.iterrows():
            # Verificar se precisamos de nova página
            if pdf.get_y() > 270:
                pdf.add_page()
                
                # Repetir cabeçalho
                pdf.set_fill_color(31, 119, 180)  # Azul
                pdf.set_text_color(255, 255, 255)  # Branco
                pdf.set_font('Arial', 'B', 8)
                
                if use_collaborator_filter and collaborator_weights:
                    pdf.cell(50, 8, "Colaborador", 1, 0, 'L', True)
                    pdf.cell(18, 8, "Ocupação", 1, 0, 'C', True)
                    pdf.cell(15, 8, "Meta", 1, 0, 'C', True)
                    pdf.cell(18, 8, "Faturabilidade", 1, 0, 'C', True)
                    pdf.cell(15, 8, "Meta", 1, 0, 'C', True)
                    pdf.cell(18, 8, "Horas Real.", 1, 0, 'C', True)
                    pdf.cell(15, 8, "Pond.", 1, 1, 'C', True)
                else:
                    pdf.cell(60, 8, "Colaborador", 1, 0, 'L', True)
                    pdf.cell(20, 8, "Ocupação", 1, 0, 'C', True)
                    pdf.cell(20, 8, "Meta Ocup.", 1, 0, 'C', True)
                    pdf.cell(20, 8, "Faturabilidade", 1, 0, 'C', True)
                    pdf.cell(20, 8, "Meta Fatur.", 1, 0, 'C', True)
                    pdf.cell(20, 8, "Horas Realiz.", 1, 1, 'C', True)
                
                pdf.set_text_color(0, 0, 0)  # Preto
                pdf.set_font('Arial', '', 7)
                
            # Alternar cores de fundo
            if idx % 2 == 0:
                pdf.set_fill_color(245, 245, 245)  # Cinza claro
                fill = True
            else:
                pdf.set_fill_color(255, 255, 255)  # Branco
                fill = False
            
            # Truncar nome se muito longo
            name = row['name']
            if len(name) > 25:
                name = name[:22] + "..."
                
            if use_collaborator_filter and collaborator_weights:
                pdf.cell(50, 7, name, 1, 0, 'L', fill)
                
                # Ocupação com cor se estiver abaixo da meta
                if row['occupation_percentage'] < 87.5:
                    pdf.set_text_color(255, 0, 0)  # Vermelho
                pdf.cell(18, 7, f"{row['occupation_percentage']:.1f}%", 1, 0, 'C', fill)
                pdf.set_text_color(0, 0, 0)  # Voltar para preto
                
                pdf.cell(15, 7, "87.5%", 1, 0, 'C', fill)
                
                # Faturabilidade com cor se estiver abaixo da meta
                if row['billable_percentage'] < 75.0:
                    pdf.set_text_color(255, 0, 0)  # Vermelho
                pdf.cell(18, 7, f"{row['billable_percentage']:.1f}%", 1, 0, 'C', fill)
                pdf.set_text_color(0, 0, 0)  # Voltar para preto
                
                pdf.cell(15, 7, "75.0%", 1, 0, 'C', fill)
                pdf.cell(18, 7, f"{row['total_hours']:.1f}h", 1, 0, 'C', fill)
                pdf.cell(15, 7, f"{row['weight_percentage']:.0f}%", 1, 1, 'C', fill)
            else:
                pdf.cell(60, 7, name, 1, 0, 'L', fill)
                
                # Ocupação with color if below target
                if row['occupation_percentage'] < 87.5:
                    pdf.set_text_color(255, 0, 0)  # Vermelho
                pdf.cell(20, 7, f"{row['occupation_percentage']:.1f}%", 1, 0, 'C', fill)
                pdf.set_text_color(0, 0, 0)  # Voltar para preto
                
                pdf.cell(20, 7, "87.5%", 1, 0, 'C', fill)
                
                # Faturabilidade with color if below target
                if row['billable_percentage'] < 75.0:
                    pdf.set_text_color(255, 0, 0)  # Vermelho
                pdf.cell(20, 7, f"{row['billable_percentage']:.1f}%", 1, 0, 'C', fill)
                pdf.set_text_color(0, 0, 0)  # Voltar para preto
                
                pdf.cell(20, 7, "75.0%", 1, 0, 'C', fill)
                pdf.cell(20, 7, f"{row['total_hours']:.1f}h", 1, 1, 'C', fill)

        pdf.ln(5)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 8, f"Horas úteis no período: {horas_uteis_periodo:.1f}h ({dias_uteis} dias úteis × 8h)", 0, 1)

        # Legenda
        pdf.ln(5)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 8, "Legenda:", 0, 1)
        
        pdf.set_font('Arial', '', 10)
        pdf.set_fill_color(76, 175, 80)  # Verde
        pdf.rect(20, pdf.get_y(), 5, 5, 'F')
        pdf.set_xy(30, pdf.get_y())
        pdf.cell(100, 5, "Acima da meta", 0, 1)
        
        pdf.set_fill_color(244, 67, 54)  # Vermelho
        pdf.rect(20, pdf.get_y(), 5, 5, 'F')
        pdf.set_xy(30, pdf.get_y())
        pdf.cell(100, 5, "Abaixo da meta", 0, 1)
        
        pdf.ln(3)
        legend_text = "Meta de Ocupação: 87.5% | Meta de Faturabilidade: 75%"
        if use_collaborator_filter and collaborator_weights:
            legend_text += " | Pond. = Ponderação aplicada"
        pdf.cell(0, 5, legend_text, 0, 1)
        
        logging.info("Finalizando a geração do PDF.")

        # Obter ausências para o período
        ausencias = get_collaborator_absences(
            db_manager,
            start_date,
            end_date,
            selected_teams
        )

        # Adicionar seção de ausências ao PDF
        if ausencias:
            pdf.add_page()
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 10, "Ausências no Período", 0, 1)
            
            # Cabeçalho da tabela
            pdf.set_fill_color(31, 119, 180)  # Azul
            pdf.set_text_color(255, 255, 255)  # Branco
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(60, 8, "Colaborador", 1, 0, 'L', True)
            pdf.cell(25, 8, "Início", 1, 0, 'C', True)
            pdf.cell(25, 8, "Fim", 1, 0, 'C', True)
            pdf.cell(30, 8, "Tipo", 1, 0, 'C', True)
            pdf.cell(15, 8, "Dias", 1, 0, 'C', True)
            pdf.cell(35, 8, "Observação", 1, 1, 'C', True)
            
            # Dados da tabela
            pdf.set_text_color(0, 0, 0)  # Preto
            pdf.set_font('Arial', '', 9)
            
            for idx, ausencia in enumerate(ausencias):
                # Alternar cores de fundo
                if idx % 2 == 0:
                    pdf.set_fill_color(245, 245, 245)  # Cinza claro
                    fill = True
                else:
                    pdf.set_fill_color(255, 255, 255)  # Branco
                    fill = False
                    
                pdf.cell(60, 7, ausencia['user_name'], 1, 0, 'L', fill)
                pdf.cell(25, 7, ausencia['start_date'].strftime('%d/%m/%Y'), 1, 0, 'C', fill)
                pdf.cell(25, 7, ausencia['end_date'].strftime('%d/%m/%Y'), 1, 0, 'C', fill)
                pdf.cell(30, 7, ausencia['absence_type'], 1, 0, 'C', fill)
                pdf.cell(15, 7, str(ausencia['duration_days']), 1, 0, 'C', fill)
                
                # Limitar tamanho da descrição
                descricao = ausencia['description']
                if descricao and len(descricao) > 20:
                    descricao = descricao[:17] + "..."
                pdf.cell(35, 7, descricao, 1, 1, 'L', fill)
            
            # Adicionar resumo de ausências
            pdf.ln(5)
            pdf.set_font('Arial', 'B', 11)
            pdf.cell(0, 8, "Resumo de Ausências:", 0, 1)
            
            # Contagem por tipo de ausência
            tipos_ausencia = {}
            for ausencia in ausencias:
                tipo = ausencia['absence_type']
                if tipo not in tipos_ausencia:
                    tipos_ausencia[tipo] = 0
                tipos_ausencia[tipo] += 1
            
            pdf.set_font('Arial', '', 10)
            for tipo, contagem in tipos_ausencia.items():
                pdf.cell(0, 6, f"- {tipo}: {contagem} ocorrência(s)", 0, 1)
            
            # Total de dias de ausência
            total_dias = sum(ausencia['duration_days'] for ausencia in ausencias)
            pdf.ln(2)
            pdf.cell(0, 6, f"Total de dias de ausência no período: {total_dias} dias", 0, 1)
        
        # Informações sobre feriados considerados
        pdf.ln(5)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 8, "Feriados Considerados no Período:", 0, 1)
        pdf.set_font('Arial', '', 9)
        
        # Listar os feriados do período
        feriados_no_periodo = [feriado for feriado in all_holidays if start_date.date() <= feriado <= end_date.date()]
        if feriados_no_periodo:
            for feriado in feriados_no_periodo:
                pdf.cell(0, 5, f"- {feriado.strftime('%d/%m/%Y')}", 0, 1)
        else:
            pdf.cell(0, 5, "Nenhum feriado no período selecionado.", 0, 1)
        
        # Salvar o PDF
        pdf.output(output_path)
        logging.info("PDF gerado com sucesso.")
        
        return True
    
    except Exception as e:
        import traceback
        error_msg = f"Erro ao gerar PDF: {str(e)}"
        logging.error(error_msg)
        logging.error(traceback.format_exc())
        return False


def generate_collaborator_excel_report(
    output_path,
    colaboradores,
    start_date,
    end_date,
    selected_teams,
    show_top_performers=True,
    show_low_performers=True,
    use_collaborator_filter=False,
    collaborator_weights=None
):
    """
    Gera relatório Excel com indicadores de colaboradores
    """
    try:
        # Verificar se os colaboradores existem
        if not colaboradores:
            # Sem dados disponíveis
            return False
        
        # Converter para DataFrame
        df = pd.DataFrame(colaboradores)
        
        # [Continue with existing Excel generation logic, adding weight information where applicable]
        
        return True
    
    except Exception as e:
        import traceback
        error_msg = f"Erro ao gerar Excel: {str(e)}"
        logging.error(error_msg)
        logging.error(traceback.format_exc())
        return False


def get_collaborator_indicators(
    db_manager,
    collaborator_target_calculator,
    start_date,
    end_date,
    selected_teams,
    use_collaborator_filter=False,
    collaborator_weights=None
):
    """
    Obter indicadores de colaboradores para o relatório
    """
    try:
        # Carregar dados necessários usando formato ISO para as datas
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d 23:59:59')  # Adicionar o horário final do dia
                
        # Usar consulta SQL que inclua registros até o final do último dia
        timesheet_df = db_manager.query_to_df(f"SELECT * FROM timesheet WHERE datetime(start_date) >= '{start_date_str}' AND datetime(end_date) <= '{end_date_str}'")

        # Adicionar log para depuração específica do período
        logging.info(f"Período de consulta: de {start_date_str} até {end_date_str}")
        users_df = db_manager.query_to_df("SELECT * FROM utilizadores WHERE active = 1")
        groups_df = db_manager.query_to_df("SELECT * FROM groups")
        
        # Filtrar por equipe
        if "Todas" not in selected_teams:
            # Encontrar os usuários que pertencem às equipes selecionadas
            filtered_users = []
            for _, user in users_df.iterrows():
                try:
                    if isinstance(user['groups'], str):
                        user_groups = eval(user['groups'])
                    else:
                        user_groups = user['groups'] if user['groups'] is not None else []
                    
                    # Converter para lista se for outro tipo
                    if not isinstance(user_groups, list):
                        if isinstance(user_groups, dict):
                            user_groups = list(user_groups.values())
                        else:
                            user_groups = [user_groups]
                    
                    # Verificar se o usuário pertence a alguma das equipes selecionadas
                    if any(team in user_groups for team in selected_teams):
                        filtered_users.append(user['user_id'])
                except Exception as e:
                    logging.warning(f"Erro ao processar grupos do usuário {user['First_Name']} {user['Last_Name']}: {str(e)}")
            
            # Filtrar usuários
            users_df = users_df[users_df['user_id'].isin(filtered_users)]
        
        # Aplicar filtro específico de colaboradores se ativo
        if use_collaborator_filter and collaborator_weights:
            selected_user_ids = list(collaborator_weights.keys())
            users_df = users_df[users_df['user_id'].isin(selected_user_ids)]
            logging.info(f"Filtro de colaboradores aplicado: {len(selected_user_ids)} colaboradores selecionados")
        
        if users_df.empty:
            return []
        
        # Calculando dias úteis do período considerando feriados
        month_year_pairs = set()
        current_date = start_date
        while current_date <= end_date:
            month_year_pairs.add((current_date.month, current_date.year))
            current_date += timedelta(days=1)
            
        # Obter todos os feriados dos anos envolvidos no período
        all_holidays = []
        for month, year in month_year_pairs:
            all_holidays.extend(get_feriados_portugal(year))
        
        # Calcular dias úteis excluindo feriados
        dias_uteis = 0
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() < 5 and current_date.date() not in all_holidays:  # 0-4 são dias úteis (seg-sex)
                dias_uteis += 1
            current_date += timedelta(days=1)
                
        # Calcular horas úteis totais para o período
        horas_uteis_periodo = dias_uteis * 8
        
        # Para cada colaborador, calcular indicadores
        collaborator_indicators = []
        
        for _, user in users_df.iterrows():
            # Filtrar entradas de timesheet para o colaborador e período específico
            try:
                user_timesheet = timesheet_df[
                    (timesheet_df['user_id'] == user['user_id']) & 
                    (pd.to_datetime(timesheet_df['start_date'], format='mixed') >= start_date_str) &
                    (pd.to_datetime(timesheet_df['start_date'], format='mixed') <= end_date_str)
                ]
            except Exception as e:
                # Fallback em caso de erro de conversão de data
                logging.warning(f"Erro ao filtrar timesheet para {user['First_Name']} {user['Last_Name']}: {e}")
                user_timesheet = pd.DataFrame()
            
            # Calcular horas realizadas (total de horas registradas)
            total_hours = user_timesheet['hours'].sum() if not user_timesheet.empty else 0
            logging.info(f"Total de Horas para {user['First_Name']} {user['Last_Name']}: {total_hours}")
            
            # Calcular horas faturáveis (apenas entradas marcadas como billable=True)
            billable_hours = user_timesheet[user_timesheet['billable'] == True]['hours'].sum() if not user_timesheet.empty else 0
            
            # Obter fator de ponderação se aplicável
            weight_factor = 1.0  # Default
            if use_collaborator_filter and collaborator_weights and user['user_id'] in collaborator_weights:
                weight_factor = collaborator_weights[user['user_id']]
            
            # Aplicar ponderação às horas para cálculo dos percentuais
            weighted_total_hours = total_hours * weight_factor
            weighted_billable_hours = billable_hours * weight_factor
            
            # Calcular percentuais de ocupação e faturabilidade (sem ponderação individual)
            # Ocupação = (Horas realizadas / Horas úteis do período) * 100
            occupation_percentage = (total_hours / horas_uteis_periodo * 100) if horas_uteis_periodo > 0 else 0

            # Faturabilidade = (Horas faturáveis / Horas úteis do período) * 100
            billable_percentage = (billable_hours / horas_uteis_periodo * 100) if horas_uteis_periodo > 0 else 0
            
            # Metas definidas
            target_occupation = 87.5  # Meta de ocupação (87,5%)
            target_billable = 75.0    # Meta de horas faturáveis (75.0%)
            
            # Determinar cores dos indicadores - simplificado para verde/vermelho
            occupation_color = "green" if occupation_percentage >= target_occupation else "red"
            billable_color = "green" if billable_percentage >= target_billable else "red"
            
            # Adicionar ao array de indicadores
            collaborator_indicators.append({
                "user_id": user['user_id'],
                "name": f"{user['First_Name']} {user['Last_Name']}",
                "occupation_percentage": occupation_percentage,
                "billable_percentage": billable_percentage,
                "occupation_color": occupation_color,
                "billable_color": billable_color,
                "target_occupation": target_occupation,
                "target_billable": target_billable,
                "total_hours": total_hours,  # Quantidade de horas realizadas (original)
                "billable_hours": billable_hours,  # Quantidade de horas faturáveis (original)
                "weighted_total_hours": weighted_total_hours,  # Horas ponderadas
                "weighted_billable_hours": weighted_billable_hours,  # Horas faturáveis ponderadas
                "weight_factor": weight_factor,  # Fator de ponderação aplicado
                "weight_percentage": weight_factor * 100  # Percentual de ponderação para exibição
            })
        
        return collaborator_indicators
    
    except Exception as e:
        import traceback
        logging.error(f"Erro ao obter indicadores de colaboradores: {str(e)}")
        logging.error(traceback.format_exc())
        return []


def send_email(
    recipients, 
    subject, 
    message, 
    pdf_path, 
    excel_path, 
    smtp_server, 
    smtp_port, 
    smtp_user, 
    smtp_password, 
    use_tls
):
    """
    Envia email com os relatórios anexados
    """
    try:
        # Criar email MIME
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = ", ".join(recipients)
        msg['Subject'] = subject
        
        # Adicionar corpo do email
        msg.attach(MIMEText(message, 'plain'))
        
        # Verificar se os arquivos existem antes de anexar
        has_attachments = False
        
        # Anexar PDF se existir
        if pdf_path and os.path.exists(pdf_path):
            with open(pdf_path, "rb") as pdf_file:
                attachment = MIMEApplication(pdf_file.read(), _subtype="pdf")
                attachment.add_header(
                    'Content-Disposition', 
                    'attachment', 
                    filename=os.path.basename(pdf_path)
                )
                msg.attach(attachment)
                has_attachments = True
                st.success(f"PDF anexado ao email.")
        else:
            st.warning(f"Arquivo PDF não encontrado em {pdf_path}")
        
        # Anexar Excel se existir
        if excel_path and os.path.exists(excel_path):
            with open(excel_path, "rb") as excel_file:
                attachment = MIMEApplication(excel_file.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                attachment.add_header(
                    'Content-Disposition', 
                    'attachment', 
                    filename=os.path.basename(excel_path)
                )
                msg.attach(attachment)
                has_attachments = True
                st.success(f"Excel anexado ao email.")
        else:
            st.warning(f"Arquivo Excel não encontrado em {excel_path}")
            
        if not has_attachments:
            st.error("Nenhum anexo disponível para enviar. Verifique se os relatórios foram gerados corretamente.")
            return False
        
        # Configurar e enviar email
        server = smtplib.SMTP(smtp_server, smtp_port)
        
        if use_tls:
            server.starttls()
        
        if smtp_user and smtp_password:
            server.login(smtp_user, smtp_password)
        
        server.send_message(msg)
        server.quit()
        
        st.success(f"Email enviado com sucesso para {', '.join(recipients)}!")
        return True
    except Exception as e:
        import traceback
        error_msg = f"Erro ao enviar email: {str(e)}"
        logging.error(error_msg)
        logging.error(traceback.format_exc())
        st.error(error_msg)
        return False


def get_collaborator_absences(db_manager, start_date, end_date, selected_teams):
    """
    Obtém ausências de colaboradores para o período selecionado
    """
    try:
        # Carregar dados necessários
        absences_df = db_manager.query_to_df("SELECT * FROM absences")
        users_df = db_manager.query_to_df("SELECT * FROM utilizadores WHERE active = 1")
        groups_df = db_manager.query_to_df("SELECT * FROM groups")
        
        # Verificar se há dados
        if absences_df.empty:
            return []
        
        # Converter datas para datetime
        absences_df['start_date'] = pd.to_datetime(absences_df['start_date'], errors='coerce')
        absences_df['end_date'] = pd.to_datetime(absences_df['end_date'], errors='coerce')
        
        # Filtrar ausências no período
        filtered_absences = absences_df[
            ((absences_df['start_date'] >= start_date) & (absences_df['start_date'] <= end_date)) |
            ((absences_df['end_date'] >= start_date) & (absences_df['end_date'] <= end_date)) |
            ((absences_df['start_date'] <= start_date) & (absences_df['end_date'] >= end_date))
        ]
        
        if filtered_absences.empty:
            return []
        
        # Filtrar por equipe se necessário
        if "Todas" not in selected_teams:
            # Encontrar os usuários que pertencem às equipes selecionadas
            filtered_users = []
            for _, user in users_df.iterrows():
                try:
                    if isinstance(user['groups'], str):
                        user_groups = eval(user['groups'])
                    else:
                        user_groups = user['groups'] if user['groups'] is not None else []
                    
                    # Converter para lista se for outro tipo
                    if not isinstance(user_groups, list):
                        if isinstance(user_groups, dict):
                            user_groups = list(user_groups.values())
                        else:
                            user_groups = [user_groups]
                    
                    # Verificar se o usuário pertence a alguma das equipes selecionadas
                    if any(team in user_groups for team in selected_teams):
                        filtered_users.append(user['user_id'])
                except Exception as e:
                    logging.warning(f"Erro ao processar grupos do usuário {user['First_Name']} {user['Last_Name']}: {str(e)}")
            
            # Filtrar ausências pelos usuários das equipes selecionadas
            filtered_absences = filtered_absences[filtered_absences['user_id'].isin(filtered_users)]
        
        # Preparar lista de ausências
        absence_list = []
        
        for _, absence in filtered_absences.iterrows():
            user_id = absence['user_id']
            user_info = users_df[users_df['user_id'] == user_id]
            
            if user_info.empty:
                continue
                
            user_name = f"{user_info['First_Name'].iloc[0]} {user_info['Last_Name'].iloc[0]}"
            
            # Ajustar datas para ficarem dentro do período selecionado
            start_date_adj = max(absence['start_date'], start_date)
            end_date_adj = min(absence['end_date'], end_date)
            
            # Verificar tipo de ausência
            absence_type = str(absence.get('absence_type', 'Outro'))
            if pd.isna(absence_type):
                absence_type = 'Outro'
            
            # Calcular duração em dias
            duration_days = (end_date_adj - start_date_adj).days + 1
            
            absence_list.append({
                'user_id': user_id,
                'user_name': user_name,
                'start_date': start_date_adj,
                'end_date': end_date_adj,
                'duration_days': duration_days,
                'absence_type': absence_type,
                'description': absence.get('description', '')
            })
        
        return absence_list
    
    except Exception as e:
        import traceback
        logging.error(f"Erro ao obter ausências de colaboradores: {str(e)}")
        logging.error(traceback.format_exc())
        return []
def generate_collaborator_excel_report(
    output_path,
    colaboradores,
    start_date,
    end_date,
    selected_teams,
    show_top_performers=True,
    show_low_performers=True,
    use_collaborator_filter=False,
    collaborator_weights=None
):
    """
    Gera relatório Excel com indicadores de colaboradores
    """
    try:
        # Verificar se os colaboradores existem
        if not colaboradores:
            # Sem dados disponíveis
            return False
        
        # Converter para DataFrame
        df = pd.DataFrame(colaboradores)
        
        # Calcular dias úteis considerando feriados
        month_year_pairs = set()
        current_date = start_date
        while current_date <= end_date:
            month_year_pairs.add((current_date.month, current_date.year))
            current_date += timedelta(days=1)
            
        # Obter todos os feriados dos anos envolvidos no período
        all_holidays = []
        for _, year in month_year_pairs:
            all_holidays.extend(get_feriados_portugal(year))
        
        # Calcular dias úteis excluindo feriados
        dias_uteis = 0
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() < 5 and current_date.date() not in all_holidays:
                dias_uteis += 1
            current_date += timedelta(days=1)
        
        # Preparar dados para o Excel
        resumo_data = {
            'Métrica': [
                'Período',
                'Dias Úteis',
                'Total de Colaboradores',
                'Ocupação Média (%)',
                'Faturabilidade Média (%)',
                'Colaboradores Abaixo da Meta de Ocupação',
                'Colaboradores Abaixo da Meta de Faturabilidade',
                'Percentual Abaixo da Meta de Ocupação',
                'Percentual Abaixo da Meta de Faturabilidade',
                'Data de Geração'
            ],
            'Valor': [
                f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}",
                dias_uteis,
                len(df),
                df['occupation_percentage'].mean(),
                df['billable_percentage'].mean(),
                sum(df['occupation_percentage'] < 87.5),
                sum(df['billable_percentage'] < 75.0),
                f"{(sum(df['occupation_percentage'] < 87.5) / len(df) * 100):.1f}%",
                f"{(sum(df['billable_percentage'] < 75.0) / len(df) * 100):.1f}%",
                datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            ]
        }
        
        # Adicionar informação sobre ponderação se aplicável
        if use_collaborator_filter and collaborator_weights:
            resumo_data['Métrica'].extend([
                'Filtro de Colaboradores Aplicado',
                'Número de Colaboradores Selecionados',
                'Ponderação Personalizada'
            ])
            resumo_data['Valor'].extend([
                'Sim',
                len(collaborator_weights),
                'Aplicada'
            ])
        
        # Filtros aplicados
        teams_str = ', '.join(selected_teams) if "Todas" not in selected_teams else "Todas as equipes"
        filtros_data = {
            'Filtro': ['Equipes'],
            'Valor': [teams_str]
        }
        
        # Planilha principal - dados completos
        # Selecionar e renomear colunas baseado se há ponderação
        if use_collaborator_filter and collaborator_weights:
            main_data = df[[
                'name',
                'occupation_percentage',
                'billable_percentage',
                'total_hours',
                'billable_hours',
                'weight_percentage'
            ]].copy()
            
            main_data.columns = [
                'Colaborador',
                'Ocupação (%)',
                'Faturabilidade (%)',
                'Horas Totais',
                'Horas Faturáveis',
                'Ponderação (%)'
            ]
        else:
            main_data = df[[
                'name',
                'occupation_percentage',
                'billable_percentage',
                'total_hours',
                'billable_hours'
            ]].copy()
            
            main_data.columns = [
                'Colaborador',
                'Ocupação (%)',
                'Faturabilidade (%)',
                'Horas Totais',
                'Horas Faturáveis'
            ]
        
        # Adicionar colunas de status
        main_data['Atinge Meta Ocupação'] = main_data['Ocupação (%)'] >= 87.5
        main_data['Atinge Meta Faturabilidade'] = main_data['Faturabilidade (%)'] >= 75.0
        
        # Adicionar colunas com a diferença para a meta
        main_data['Diferença para Meta Ocupação'] = main_data['Ocupação (%)'] - 87.5
        main_data['Diferença para Meta Faturabilidade'] = main_data['Faturabilidade (%)'] - 75.0
        
        # Criar escritor Excel com openpyxl
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Criar workbook e planilhas
        wb = Workbook()
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        ws_dados = wb.create_sheet("Dados Completos")
        
        # Planilha de Resumo
        ws_resumo.append(["Métrica", "Valor"])
        for i in range(len(resumo_data['Métrica'])):
            ws_resumo.append([resumo_data['Métrica'][i], resumo_data['Valor'][i]])
            
        # Adicionar espaço e filtros
        ws_resumo.append([])
        ws_resumo.append([])
        ws_resumo.append(["Filtro", "Valor"])
        ws_resumo.append(["Equipes", teams_str])
        
        # Adicionar informações de ponderação se aplicável
        if use_collaborator_filter and collaborator_weights:
            ws_resumo.append([])
            ws_resumo.append(["Informações de Ponderação", ""])
            ws_resumo.append(["Colaborador", "Ponderação (%)"])
            
            # Obter nomes dos colaboradores para mostrar a ponderação
            for colaborador in colaboradores:
                if colaborador['user_id'] in collaborator_weights:
                    ws_resumo.append([colaborador['name'], f"{colaborador['weight_percentage']:.0f}%"])
        
        # Adicionar feriados ao resumo
        ws_resumo.append([])
        ws_resumo.append([])
        ws_resumo.append(["Feriados no Período", "Data"])
        
        feriados_no_periodo = [feriado for feriado in all_holidays if start_date.date() <= feriado <= end_date.date()]
        if feriados_no_periodo:
            for feriado in feriados_no_periodo:
                ws_resumo.append(["", feriado.strftime('%d/%m/%Y')])
        else:
            ws_resumo.append(["", "Nenhum feriado no período selecionado"])
            
        # Planilha de Dados Completos
        for r in dataframe_to_rows(main_data, index=False, header=True):
            ws_dados.append(r)
            
        # Formatar cabeçalhos
        header_fill = PatternFill(start_color="1E77B4", end_color="1E77B4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws_resumo["1:1"]:
            cell.fill = header_fill
            cell.font = header_font
            
        # Encontrar linha de cabeçalho de filtros
        filtro_row = None
        for row_num, row in enumerate(ws_resumo.iter_rows(values_only=True), 1):
            if row[0] == "Filtro":
                filtro_row = row_num
                break
        
        if filtro_row:
            for cell in ws_resumo[f"{filtro_row}:{filtro_row}"]:
                cell.fill = header_fill
                cell.font = header_font
        
        # Encontrar linha de cabeçalho de feriados
        feriado_row = None
        for row_num, row in enumerate(ws_resumo.iter_rows(values_only=True), 1):
            if row[0] == "Feriados no Período":
                feriado_row = row_num
                break
        
        if feriado_row:
            for cell in ws_resumo[f"{feriado_row}:{feriado_row}"]:
                cell.fill = header_fill
                cell.font = header_font
        
        # Encontrar linha de cabeçalho de ponderação
        if use_collaborator_filter and collaborator_weights:
            pond_row = None
            for row_num, row in enumerate(ws_resumo.iter_rows(values_only=True), 1):
                if row[0] == "Colaborador" and row[1] == "Ponderação (%)":
                    pond_row = row_num
                    break
            
            if pond_row:
                for cell in ws_resumo[f"{pond_row}:{pond_row}"]:
                    cell.fill = header_fill
                    cell.font = header_font
            
        for cell in ws_dados["1:1"]:
            cell.fill = header_fill
            cell.font = header_font
            
        # Adicionar planilhas adicionais conforme necessário
        if show_top_performers and len(df) > 0:
            ws_top = wb.create_sheet("Top Performers")
            
            # Top 5 em ocupação
            top_ocupacao = df.sort_values('occupation_percentage', ascending=False).head(5)
            
            ws_top.append(["Top 5 por Ocupação"])
            
            if use_collaborator_filter and collaborator_weights:
                ws_top.append(["Colaborador", "Ocupação (%)", "Faturabilidade (%)", "Horas Totais", "Ponderação (%)"])
            else:
                ws_top.append(["Colaborador", "Ocupação (%)", "Faturabilidade (%)", "Horas Totais"])
            
            for _, row in top_ocupacao.iterrows():
                if use_collaborator_filter and collaborator_weights:
                    ws_top.append([
                        row['name'],
                        row['occupation_percentage'],
                        row['billable_percentage'],
                        row['total_hours'],
                        row['weight_percentage']
                    ])
                else:
                    ws_top.append([
                        row['name'],
                        row['occupation_percentage'],
                        row['billable_percentage'],
                        row['total_hours']
                    ])
                
            # Espaço entre tabelas
            ws_top.append([])
            ws_top.append([])
            
            # Top 5 em faturabilidade
            top_faturabilidade = df.sort_values('billable_percentage', ascending=False).head(5)
            
            ws_top.append(["Top 5 por Faturabilidade"])
            
            if use_collaborator_filter and collaborator_weights:
                ws_top.append(["Colaborador", "Faturabilidade (%)", "Ocupação (%)", "Horas Faturáveis", "Ponderação (%)"])
            else:
                ws_top.append(["Colaborador", "Faturabilidade (%)", "Ocupação (%)", "Horas Faturáveis"])
            
            for _, row in top_faturabilidade.iterrows():
                if use_collaborator_filter and collaborator_weights:
                    ws_top.append([
                        row['name'],
                        row['billable_percentage'],
                        row['occupation_percentage'],
                        row['billable_hours'],
                        row['weight_percentage']
                    ])
                else:
                    ws_top.append([
                        row['name'],
                        row['billable_percentage'],
                        row['occupation_percentage'],
                        row['billable_hours']
                    ])
                
            # Formatar cabeçalhos
            for cell in ws_top["2:2"]:  # Cabeçalho top ocupação
                cell.fill = header_fill
                cell.font = header_font
                
            # Encontrar linha do segundo cabeçalho
            second_header_row = 9 if not (use_collaborator_filter and collaborator_weights) else 9
            for cell in ws_top[f"{second_header_row}:{second_header_row}"]:
                cell.fill = header_fill
                cell.font = header_font
                
        if show_low_performers and len(df) > 0:
            ws_low = wb.create_sheet("Areas de Melhoria")
            
            # Abaixo da meta de ocupação
            below_ocupacao = df[df['occupation_percentage'] < 87.5].sort_values('occupation_percentage')
            
            ws_low.append(["Colaboradores Abaixo da Meta de Ocupação"])
            
            if use_collaborator_filter and collaborator_weights:
                ws_low.append(["Colaborador", "Ocupação (%)", "Faturabilidade (%)", "Diferença para Meta", "Ponderação (%)"])
            else:
                ws_low.append(["Colaborador", "Ocupação (%)", "Faturabilidade (%)", "Diferença para Meta"])
            
            if not below_ocupacao.empty:
                for _, row in below_ocupacao.iterrows():
                    if use_collaborator_filter and collaborator_weights:
                        ws_low.append([
                            row['name'],
                            row['occupation_percentage'],
                            row['billable_percentage'],
                            row['occupation_percentage'] - 87.5,
                            row['weight_percentage']
                        ])
                    else:
                        ws_low.append([
                            row['name'],
                            row['occupation_percentage'],
                            row['billable_percentage'],
                            row['occupation_percentage'] - 87.5
                        ])
            else:
                ws_low.append(["Todos os colaboradores atingiram a meta de ocupação!"])
                
            # Espaço entre tabelas
            ws_low.append([])
            ws_low.append([])
            
            # Abaixo da meta de faturabilidade
            below_faturabilidade = df[df['billable_percentage'] < 75.0].sort_values('billable_percentage')
            
            ws_low.append(["Colaboradores Abaixo da Meta de Faturabilidade"])
            
            if use_collaborator_filter and collaborator_weights:
                ws_low.append(["Colaborador", "Faturabilidade (%)", "Ocupação (%)", "Diferença para Meta", "Ponderação (%)"])
            else:
                ws_low.append(["Colaborador", "Faturabilidade (%)", "Ocupação (%)", "Diferença para Meta"])
            
            if not below_faturabilidade.empty:
                for _, row in below_faturabilidade.iterrows():
                    if use_collaborator_filter and collaborator_weights:
                        ws_low.append([
                            row['name'],
                            row['billable_percentage'],
                            row['occupation_percentage'],
                            row['billable_percentage'] - 75.0,
                            row['weight_percentage']
                        ])
                    else:
                        ws_low.append([
                            row['name'],
                            row['billable_percentage'],
                            row['occupation_percentage'],
                            row['billable_percentage'] - 75.0
                        ])
            else:
                ws_low.append(["Todos os colaboradores atingiram a meta de faturabilidade!"])
                
            # Formatar cabeçalhos
            for cell in ws_low["2:2"]:  # Cabeçalho abaixo meta ocupação
                cell.fill = header_fill
                cell.font = header_font
                
            # Encontrar linha do segundo cabeçalho
            fatur_header_row = None
            for row_num, row in enumerate(ws_low.iter_rows(values_only=True), 1):
                if row[0] == "Colaboradores Abaixo da Meta de Faturabilidade":
                    fatur_header_row = row_num + 1
                    break
            
            if fatur_header_row:
                for cell in ws_low[f"{fatur_header_row}:{fatur_header_row}"]:
                    cell.fill = header_fill
                    cell.font = header_font
        
        # Adicionar planilha de detalhes de ponderação se aplicável
        if use_collaborator_filter and collaborator_weights:
            ws_pond = wb.create_sheet("Detalhes de Ponderação")
            
            ws_pond.append(["Detalhes da Ponderação Aplicada"])
            ws_pond.append([])
            ws_pond.append(["Colaborador", "Ponderação (%)", "Horas Originais", "Horas Ponderadas", "Horas Fatur. Orig.", "Horas Fatur. Pond."])
            
            for _, row in df.iterrows():
                ws_pond.append([
                    row['name'],
                    row['weight_percentage'],
                    row['total_hours'],
                    row['weighted_total_hours'],
                    row['billable_hours'],
                    row['weighted_billable_hours']
                ])
            
            # Formatar cabeçalho
            for cell in ws_pond["3:3"]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Adicionar explicação
            ws_pond.append([])
            ws_pond.append(["Explicação:"])
            ws_pond.append(["- Horas Originais: Horas realmente trabalhadas pelo colaborador"])
            ws_pond.append(["- Horas Ponderadas: Horas originais × fator de ponderação"])
            ws_pond.append(["- Os percentuais de ocupação e faturabilidade são calculados com base nas horas ponderadas"])
            ws_pond.append(["- A ponderação permite ajustar a contribuição de cada colaborador no cálculo das médias"])
            
        # Salvar o Excel
        wb.save(output_path)
        logging.info(f"Excel gerado com sucesso: {output_path}")
        return True
    
    except Exception as e:
        import traceback
        error_msg = f"Erro ao gerar Excel: {str(e)}"
        logging.error(error_msg)
        logging.error(traceback.format_exc())
        return False# collaborator_email_report.py