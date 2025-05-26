import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
import os
import smtplib
import tempfile
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage
from datetime import datetime, timedelta
import calendar
from fpdf import FPDF
import base64
from database_manager import DatabaseManager

def project_status_email():
    """
    Módulo para geração e envio de relatórios executivos de status de projetos por email
    """
    st.title("📧 Relatórios Executivos de Projetos")
    
    # Inicializar o gerenciador de banco de dados
    db_manager = DatabaseManager()
    
    # Verificar se o usuário é administrador
    if st.session_state.user_info['role'].lower() != 'admin':
        st.warning("Esta funcionalidade é exclusiva para administradores.")
        return
    
    # Carregamento de dados
    try:
        clients_df = db_manager.query_to_df("SELECT * FROM clients WHERE active = 1")
        projects_df = db_manager.query_to_df("SELECT * FROM projects WHERE status = 'active'")
        users_df = db_manager.query_to_df("SELECT * FROM utilizadores WHERE active = 1")
        timesheet_df = db_manager.query_to_df("SELECT * FROM timesheet")
        groups_df = db_manager.query_to_df("SELECT * FROM groups WHERE active = 1")
        rates_df = db_manager.query_to_df("SELECT * FROM rates")
        
        # Converter datas
        timesheet_df['start_date'] = pd.to_datetime(timesheet_df['start_date'], format='mixed', errors='coerce')
        projects_df['start_date'] = pd.to_datetime(projects_df['start_date'], format='mixed', errors='coerce')
        projects_df['end_date'] = pd.to_datetime(projects_df['end_date'], format='mixed', errors='coerce')
        
        # Obter tipos de projetos únicos
        project_types = sorted(projects_df['project_type'].unique().tolist())
        
    except Exception as e:
        st.error(f"Erro ao carregar dados: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return
    
    # Interface para configuração do relatório
    st.subheader("Configuração do Relatório")
    
    with st.form("email_report_config"):
        # Seleção de destinatários
        recipients = st.text_input(
            "Destinatários (separados por vírgula)",
            help="Email dos destinatários separados por vírgula"
        )
        
        # Assunto do email
        subject = st.text_input(
            "Assunto do Email", 
            f"Relatório Executivo de Projetos - {datetime.now().strftime('%d/%m/%Y')}"
        )
        
        # Filtros para o relatório
        col1, col2 = st.columns(2)
        
        with col1:
            report_period = st.selectbox(
                "Período do Relatório",
                ["Último Mês", "Últimos 3 Meses", "Últimos 6 Meses", "Ano Atual", "Período Personalizado"]
            )
            
            if report_period == "Período Personalizado":
                start_date = st.date_input(
                    "Data de Início",
                    value=datetime.now().replace(day=1) - timedelta(days=30)
                )
                end_date = st.date_input(
                    "Data de Término",
                    value=datetime.now()
                )
            else:
                # Definir datas automaticamente conforme o período selecionado
                if report_period == "Último Mês":
                    last_month = datetime.now().replace(day=1) - timedelta(days=1)
                    start_date = last_month.replace(day=1)
                    end_date = last_month.replace(day=calendar.monthrange(last_month.year, last_month.month)[1])
                elif report_period == "Últimos 3 Meses":
                    start_date = (datetime.now() - timedelta(days=90)).replace(day=1)
                    end_date = datetime.now()
                elif report_period == "Últimos 6 Meses":
                    start_date = (datetime.now() - timedelta(days=180)).replace(day=1)
                    end_date = datetime.now()
                elif report_period == "Ano Atual":
                    start_date = datetime(datetime.now().year, 1, 1)
                    end_date = datetime.now()
            
            # Novos filtros de tipos de projeto (multi-select)
            selected_project_types = st.multiselect(
                "Tipos de Projetos",
                options=["Todos"] + project_types,
                default=["Todos"],
                help="Selecione os tipos de projeto para incluir no relatório"
            )
        
        with col2:
            # Seleção de equipes (múltipla)
            team_options = ["Todas"] + sorted(groups_df['group_name'].tolist())
            selected_teams = st.multiselect(
                "Equipes", 
                options=team_options,
                default=["Todas"]
            )
            
            # Seleção de clientes (múltipla)
            client_options = ["Todos"] + sorted(clients_df['name'].tolist())
            selected_clients = st.multiselect(
                "Clientes", 
                options=client_options,
                default=["Todos"]
            )
            
            # Seleção de formato de relatório
            report_format = st.radio(
                "Formato do Relatório",
                ["PDF", "Excel", "PDF e Excel"]
            )
        
        # Conteúdo do email
        email_message = st.text_area(
            "Mensagem do Email",
            """Prezados,

Em anexo, relatório executivo atualizado com os status dos projetos.

Atenciosamente,
Equipe de Gestão de Projetos"""
        )
        
        # Opções adicionais
        include_charts = st.checkbox("Incluir Gráficos", value=True)
        include_financial = st.checkbox("Incluir Informações Financeiras", value=True)
        include_hour_details = st.checkbox("Incluir Detalhes de Horas", value=True)
        
        # SMTP settings (collapsed by default)
        with st.expander("Configurações de SMTP"):
            smtp_server = st.text_input("Servidor SMTP", "smtp.office365.com")
            smtp_port = st.number_input("Porta SMTP", value=587, step=1)
            smtp_user = st.text_input("Usuário SMTP", "notifications@grupoerre.pt")
            smtp_password = st.text_input("Senha SMTP", type="password", value="erretech@2020")
            use_tls = st.checkbox("Usar TLS", value=True)
        
        # Botão para gerar e enviar o relatório
        submit_button = st.form_submit_button("Gerar e Enviar Relatório")
    
    if submit_button:
        # Validar entradas
        if not recipients:
            st.error("Por favor, informe pelo menos um destinatário.")
            return
        
        # Filtrar dados conforme os filtros selecionados
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())
        
        # Filtrar dados de timesheet pelo período
        filtered_timesheet = timesheet_df[
            (timesheet_df['start_date'] >= start_datetime) & 
            (timesheet_df['start_date'] <= end_datetime)
        ]
        
        # Filtrar por tipos de projeto se não for "Todos"
        if "Todos" not in selected_project_types:
            filtered_projects = projects_df[projects_df['project_type'].isin(selected_project_types)]
            # Filtrar timesheet apenas para os projetos dos tipos selecionados
            project_ids = filtered_projects['project_id'].tolist()
            filtered_timesheet = filtered_timesheet[filtered_timesheet['project_id'].isin(project_ids)]
        else:
            filtered_projects = projects_df.copy()
        
        # Filtrar por equipes se não for "Todas"
        if "Todas" not in selected_teams:
            # Obter IDs das equipes selecionadas
            team_ids = groups_df[groups_df['group_name'].isin(selected_teams)]['id'].tolist()
            
            # Filtrar projetos por equipe
            filtered_projects = filtered_projects[filtered_projects['group_id'].isin(team_ids)]
            
            # Filtrar timesheet apenas para os projetos das equipes selecionadas
            project_ids = filtered_projects['project_id'].tolist()
            filtered_timesheet = filtered_timesheet[filtered_timesheet['project_id'].isin(project_ids)]
        
        # Filtrar por clientes se não for "Todos"
        if "Todos" not in selected_clients:
            # Obter IDs dos clientes selecionados
            client_ids = clients_df[clients_df['name'].isin(selected_clients)]['client_id'].tolist()
            
            # Filtrar projetos por cliente
            filtered_projects = filtered_projects[filtered_projects['client_id'].isin(client_ids)]
            
            # Filtrar timesheet apenas para os projetos dos clientes selecionados
            project_ids = filtered_projects['project_id'].tolist()
            filtered_timesheet = filtered_timesheet[filtered_timesheet['project_id'].isin(project_ids)]
        
        if filtered_projects.empty:
            st.error("Não foram encontrados projetos com os filtros selecionados.")
            return
        
        with st.spinner("Gerando e enviando relatório..."):
            # Criar diretório temporário para os arquivos
            temp_dir = tempfile.mkdtemp()
            
            # Gerar relatórios conforme formato selecionado
            pdf_path = None
            excel_path = None
            
            if "PDF" in report_format:
                pdf_path = os.path.join(temp_dir, "relatorio_status_projetos.pdf")
                pdf_result = generate_pdf_report(
                    pdf_path, 
                    filtered_projects, 
                    filtered_timesheet, 
                    clients_df, 
                    users_df,
                    rates_df,
                    groups_df,
                    start_datetime, 
                    end_datetime, 
                    include_charts, 
                    include_financial,
                    include_hour_details,
                    selected_teams,
                    selected_clients,
                    selected_project_types
                )
                
                # Verificar se o PDF foi gerado com sucesso
                if not pdf_result or not os.path.exists(pdf_path):
                    st.error(f"Falha ao gerar o PDF. Verifique os logs para mais detalhes.")
                    pdf_path = None

            if "Excel" in report_format:
                excel_path = os.path.join(temp_dir, "relatorio_status_projetos.xlsx")
                try:
                    generate_excel_report(
                        excel_path, 
                        filtered_projects, 
                        filtered_timesheet, 
                        clients_df, 
                        users_df,
                        rates_df,
                        groups_df,
                        start_datetime, 
                        end_datetime, 
                        include_financial,
                        include_hour_details,
                        selected_teams,
                        selected_clients,
                        selected_project_types
                    )
                    if os.path.exists(excel_path):
                        st.success("Excel gerado com sucesso!")
                    else:
                        st.error("Falha ao gerar o Excel.")
                        excel_path = None
                except Exception as e:
                    st.error(f"Erro ao gerar Excel: {str(e)}")
                    excel_path = None
            
            # Enviar email com os relatórios gerados se pelo menos um foi gerado com sucesso
            if pdf_path or excel_path:
                email_success = send_email(
                    recipients.split(','), 
                    subject, 
                    email_message, 
                    pdf_path, 
                    excel_path, 
                    smtp_server, 
                    smtp_port, 
                    smtp_user, 
                    smtp_password, 
                    use_tls
                )
                
                # Botões para download local dos relatórios
                st.markdown("### Download dos Relatórios")
                
                if pdf_path and os.path.exists(pdf_path):
                    with open(pdf_path, "rb") as pdf_file:
                        pdf_bytes = pdf_file.read()
                        st.download_button(
                            label="📥 Baixar PDF",
                            data=pdf_bytes,
                            file_name="relatorio_status_projetos.pdf",
                            mime="application/pdf"
                        )
                
                if excel_path and os.path.exists(excel_path):
                    with open(excel_path, "rb") as excel_file:
                        excel_bytes = excel_file.read()
                        st.download_button(
                            label="📥 Baixar Excel",
                            data=excel_bytes,
                            file_name="relatorio_status_projetos.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            else:
                st.error("Não foi possível gerar nenhum dos relatórios. Verifique os logs para mais detalhes.")
                
                # Ainda permitir download mesmo que o email falhe
                if pdf_path and os.path.exists(pdf_path):
                    with open(pdf_path, "rb") as pdf_file:
                        st.download_button(
                            label="📥 Baixar PDF",
                            data=pdf_file,
                            file_name="relatorio_status_projetos.pdf",
                            mime="application/pdf"
                        )
                
                if excel_path and os.path.exists(excel_path):
                    with open(excel_path, "rb") as excel_file:
                        st.download_button(
                            label="📥 Baixar Excel",
                            data=excel_file,
                            file_name="relatorio_status_projetos.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

class ImprovedPDF(FPDF):
    """Classe personalizada de PDF com melhor formatação e funcionalidades adicionais"""
    
    def __init__(self, orientation='P', unit='mm', format='A4'):
        super().__init__(orientation, unit, format)
        self.add_font('DejaVu', '', './fonts/DejaVuSansCondensed.ttf', uni=True)
        self.add_font('DejaVu', 'B', './fonts/DejaVuSansCondensed-Bold.ttf', uni=True)
        self.add_font('DejaVu', 'I', './fonts/DejaVuSansCondensed-Oblique.ttf', uni=True)
        
        # Definir cores personalizadas
        self.primary_color = (31, 119, 180)  # Azul
        self.secondary_color = (44, 160, 44)  # Verde
        self.accent_color = (255, 127, 14)   # Laranja
        self.warning_color = (214, 39, 40)   # Vermelho
        
        self.set_margins(20, 20, 20)
        self.set_auto_page_break(True, margin=20)
    
    def header(self):
        """Cabeçalho personalizado para todas as páginas"""
        # Logo se existir (tamanho ajustado)
        if os.path.exists('logo.png'):
            self.image('logo.png', 20, 10, 30)
        
        # Linha decorativa
        self.set_draw_color(*self.primary_color)
        self.set_line_width(0.5)
        self.line(20, 25, 190, 25)
        
        # Não mostrar cabeçalho na primeira página (será tratada separadamente)
        if self.page_no() != 1:
            # Título do relatório
            self.set_font('DejaVu', 'B', 12)
            self.set_text_color(*self.primary_color)
            self.cell(0, 10, 'Relatório Executivo de Status de Projetos', 0, 0, 'R')
    
    def footer(self):
        """Rodapé personalizado para todas as páginas"""
        # Posição a 1.5 cm do final
        self.set_y(-15)
        
        # Linha decorativa
        self.set_draw_color(*self.primary_color)
        self.set_line_width(0.3)
        self.line(20, self.get_y() - 3, 190, self.get_y() - 3)
        
        # Data de geração e número da página
        self.set_font('DejaVu', 'I', 8)
        self.set_text_color(128, 128, 128)  # Cinza
        
        # Data à esquerda
        self.cell(0, 10, f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", 0, 0, 'L')
        
        # Página à direita
        self.cell(0, 10, f'Página {self.page_no()}/{self.alias_nb_pages()}', 0, 0, 'R')
    
    def chapter_title(self, title, color=None):
        """Formata um título de capítulo"""
        if color is None:
            color = self.primary_color
            
        # Verifica se o título ficaria isolado no fim da página e força nova página se necessário
        if self.get_y() > 240:
            self.add_page()
            
        # Espaçamento superior
        self.ln(10)
        
        # Configurações do título
        self.set_font('DejaVu', 'B', 14)
        self.set_text_color(*color)
        
        # Fundo decorativo
        self.set_fill_color(240, 240, 240)
        self.cell(0, 10, title, 0, 1, 'L', True)
        
        # Linha decorativa
        self.set_draw_color(*color)
        self.set_line_width(0.3)
        self.line(20, self.get_y(), 190, self.get_y())
        
        # Espaçamento final
        self.ln(5)
        
        # Restaurar cor de texto padrão
        self.set_text_color(0, 0, 0)
    
    def section_title(self, title, color=None):
        """Formata um título de seção"""
        if color is None:
            color = self.primary_color
            
        # Verifica se o título ficaria isolado no fim da página e força nova página se necessário
        if self.get_y() > 250:
            self.add_page()
            
        # Espaçamento superior
        self.ln(5)
        
        # Configurações do título
        self.set_font('DejaVu', 'B', 12)
        self.set_text_color(*color)
        self.cell(0, 7, title, 0, 1, 'L')
        
        # Espaçamento final
        self.ln(2)
        
        # Restaurar cor de texto padrão
        self.set_text_color(0, 0, 0)
    
    def create_table_header(self, headers, widths, height=10, fill_color=None):
        """Cria um cabeçalho de tabela formatado"""
        if fill_color is None:
            fill_color = self.primary_color
            
        # Verifica se cabeçalho ficaria isolado no fim da página e força nova página se necessário
        if self.get_y() > 240:
            self.add_page()
            
        # Configurações do cabeçalho
        self.set_font('DejaVu', 'B', 10)
        self.set_text_color(255, 255, 255)  # Texto branco
        self.set_fill_color(*fill_color)
        
        # Desenhar células
        for i, header in enumerate(headers):
            self.cell(widths[i], height, header, 1, 0, 'C', True)
        self.ln()
        
        # Restaurar cor de texto padrão
        self.set_text_color(0, 0, 0)
    
    def add_table_row(self, data, widths, height=8, alternate=False, align_list=None):
        """Adiciona uma linha de tabela com formatação alternada opcional"""
        if align_list is None:
            align_list = ['L'] * len(data)
            
        # Configurações da linha
        self.set_font('DejaVu', '', 9)
        
        # Cor de fundo alternada
        if alternate:
            self.set_fill_color(240, 240, 240)
        else:
            self.set_fill_color(255, 255, 255)
        
        # Desenhar células
        for i, cell_data in enumerate(data):
            self.cell(widths[i], height, str(cell_data), 1, 0, align_list[i], alternate)
        self.ln()
    
    def add_info_box(self, title, content, color=None):
        """Adiciona uma caixa de informação estilizada"""
        if color is None:
            color = self.secondary_color
            
        # Verifica se a caixa ficaria cortada entre páginas
        if self.get_y() > 220:
            self.add_page()
            
        # Configurações da caixa
        self.set_draw_color(*color)
        self.set_fill_color(240, 240, 240)
        
        # Calcular altura total
        self.set_font('DejaVu', 'B', 11)
        title_height = 8
        
        self.set_font('DejaVu', '', 9)
        # Calcular altura aproximada para o conteúdo (texto multilinhas)
        content_lines = len(content) / 80 + content.count('\n') + 1  # Aproximação grosseira
        content_height = content_lines * 5
        
        # Desenhar retângulo de fundo
        total_height = title_height + content_height + 10  # +10 para padding
        self.rect(20, self.get_y(), 170, total_height, 'DF')
        
        # Título
        self.set_font('DejaVu', 'B', 11)
        self.set_text_color(*color)
        self.set_xy(25, self.get_y() + 5)
        self.cell(160, title_height, title, 0, 1, 'L')
        
        # Conteúdo
        self.set_text_color(0, 0, 0)
        self.set_font('DejaVu', '', 9)
        self.set_xy(25, self.get_y())
        self.multi_cell(160, 5, content, 0, 'L')
        
        # Ajustar posição Y
        self.set_y(self.get_y() + 5)

def calculate_project_metrics(projects_df, timesheet_df, clients_df, users_df, rates_df):
    """
    Calcula métricas detalhadas para cada projeto
    """
    project_metrics = {}
    
    for _, project in projects_df.iterrows():
        project_id = project['project_id']
        
        # Buscar entradas de timesheet para o projeto
        project_timesheet = timesheet_df[timesheet_df['project_id'] == project_id]
        
        # Calcular horas realizadas
        realized_hours = project_timesheet['hours'].sum() if not project_timesheet.empty else 0
        
        # Calcular custo realizado
        realized_cost = 0
        if not project_timesheet.empty:
            for _, entry in project_timesheet.iterrows():
                user_id = entry['user_id']
                hours = float(entry['hours'])
                
                # Verificar se a entrada é hora extra
                is_overtime = entry.get('overtime', False)
                if isinstance(is_overtime, (int, float)):
                    is_overtime = bool(is_overtime)
                elif isinstance(is_overtime, str):
                    is_overtime = is_overtime.lower() in ('true', 't', 'yes', 'y', '1')
                
                # Obter rate do usuário
                if not users_df.empty:
                    user_info = users_df[users_df['user_id'] == user_id]
                    if not user_info.empty and not pd.isna(user_info['rate_id'].iloc[0]):
                        rate_id = user_info['rate_id'].iloc[0]
                        rate_info = rates_df[rates_df['rate_id'] == rate_id]
                        if not rate_info.empty:
                            rate_value = float(rate_info['rate_cost'].iloc[0])
                            entry_cost = hours * rate_value
                            
                            # Se for hora extra, multiplicar por 2
                            if is_overtime:
                                entry_cost *= 2
                                
                            realized_cost += entry_cost
        
        # Total de horas e custo planejados
        total_hours = float(project['total_hours'])
        total_cost = float(project['total_cost'])
        
        # Datas do projeto
        start_date = pd.to_datetime(project['start_date'])
        end_date = pd.to_datetime(project['end_date'])
        
        # Dias totais e restantes
        total_days = (end_date - start_date).days
        days_remaining = (end_date - datetime.now()).days
        days_remaining = max(0, days_remaining)
        
        # Percentual de horas gastas (em vez de conclusão)
        hours_percentage = (realized_hours / total_hours * 100) if total_hours > 0 else 0
        
        # Percentual de custo gasto
        cost_percentage = (realized_cost / total_cost * 100) if total_cost > 0 else 0
        
        # Percentual de tempo decorrido baseado no cronograma do projeto
        elapsed_days = (datetime.now() - start_date).days
        elapsed_days = max(0, min(elapsed_days, total_days))  # Garantir que esteja entre 0 e total_days
        time_percentage = (elapsed_days / total_days * 100) if total_days > 0 else 0
        
        # Calcular CPI (Cost Performance Index) baseado no cronograma do projeto
        # CPI = (Valor Planejado / Custo Real)
        # Valor Planejado = % tempo decorrido * custo total planejado
        planned_value = (time_percentage / 100) * total_cost
        cpi = planned_value / realized_cost if realized_cost > 0 else 1.0
        
        # Determinar nível de risco e razão
        if cpi >= 1.1:
            risk_level = "Baixo"
            risk_reason = "O projeto está abaixo do orçamento para o período atual"
        elif cpi >= 0.9:
            risk_level = "Médio"
            risk_reason = "O projeto está próximo do orçamento projetado para o período atual"
        else:
            risk_level = "Alto"
            risk_reason = "O projeto está acima do orçamento para o período atual"
        
        # Adicionar análise adicional de risco baseada em desvios de cronograma
        schedule_variance = hours_percentage - time_percentage
        
        if schedule_variance > 15:  # Horas consumidas muito acima do esperado pelo tempo
            risk_level = "Alto" if risk_level != "Alto" else risk_level
            risk_reason += f". Consumo de horas está {schedule_variance:.1f}% acima do esperado para o cronograma"
        elif schedule_variance < -15:  # Horas consumidas muito abaixo do esperado
            if risk_level == "Baixo":
                risk_reason += f". Consumo de horas está {abs(schedule_variance):.1f}% abaixo do esperado para o cronograma"
            else:
                risk_level = "Médio" if risk_level == "Alto" else risk_level
                risk_reason += f". Apesar disso, o consumo de horas está {abs(schedule_variance):.1f}% abaixo do esperado"
        
        # Armazenar métricas calculadas
        project_metrics[project_id] = {
            'realized_hours': realized_hours,
            'total_hours': total_hours,
            'realized_cost': realized_cost,
            'total_cost': total_cost,
            'days_remaining': days_remaining,
            'total_days': total_days,
            'hours_percentage': hours_percentage,  # Novo: percentual de horas gastas
            'cost_percentage': cost_percentage,    # Novo: percentual de custo gasto
            'time_percentage': time_percentage,
            'cpi': cpi,
            'risk_level': risk_level,
            'risk_reason': risk_reason  # Novo: razão da classificação de risco
        }
    
    return project_metrics

def get_resource_hours(timesheet_df, project_id, users_df):
    """
    Obtém horas por recurso (colaborador) para um projeto específico
    """
    # Filtrar apenas entradas do projeto
    project_entries = timesheet_df[timesheet_df['project_id'] == project_id]
    
    if project_entries.empty:
        return []
    
    # Agrupar horas por usuário
    hours_by_user = project_entries.groupby('user_id')['hours'].sum().reset_index()
    
    # Total de horas do projeto
    total_hours = hours_by_user['hours'].sum()
    
    # Resultado final
    resource_hours = []
    
    for _, row in hours_by_user.iterrows():
        user_id = row['user_id']
        hours = row['hours']
        
        # Obter nome do usuário
        user_info = users_df[users_df['user_id'] == user_id]
        if not user_info.empty:
            first_name = user_info['First_Name'].iloc[0]
            last_name = user_info['Last_Name'].iloc[0]
            name = f"{first_name} {last_name}"
        else:
            name = f"Usuário ID: {user_id}"
        
        # Percentual das horas totais
        percentage = (hours / total_hours * 100) if total_hours > 0 else 0
        
        resource_hours.append({
            'user_id': user_id,
            'name': name,
            'hours': hours,
            'percentage': percentage
        })
    
    # Ordenar por horas (decrescente)
    resource_hours.sort(key=lambda x: x['hours'], reverse=True)
    
    return resource_hours

def generate_excel_report(
    output_path, 
    projects_df, 
    timesheet_df, 
    clients_df, 
    users_df, 
    rates_df,
    groups_df,
    start_date, 
    end_date, 
    include_financial,
    include_hour_details,
    selected_teams,
    selected_clients,
    selected_project_types
):
    """
    Gera relatório Excel com status dos projetos
    """
    # Calcular métricas para cada projeto
    project_metrics = calculate_project_metrics(
        projects_df, 
        timesheet_df, 
        clients_df, 
        users_df, 
        rates_df
    )
    
    # Criar escritor Excel
    import pandas as pd
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, colors
    
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    
    # Dados para a planilha de resumo
    resumo_data = {
        'Métrica': [
            'Total de Projetos Ativos',
            'Total de Horas Registradas',
            'Total de Horas Planejadas',
            'Percentual Médio de Horas Gastas'  # Alterado de "Conclusão" para "Horas Gastas"
        ],
        'Valor': [
            len(project_metrics),
            sum(metrics['realized_hours'] for metrics in project_metrics.values()),
            sum(metrics['total_hours'] for metrics in project_metrics.values()),
            sum(metrics['hours_percentage'] for metrics in project_metrics.values()) / len(project_metrics) if project_metrics else 0
        ]
    }
    
    # Adicionar métricas financeiras se necessário
    if include_financial:
        resumo_data['Métrica'].extend([
            'Custo Total Planejado',
            'Custo Total Realizado',
            'Percentual Médio de Custo Gasto',  # Novo campo
            'CPI Médio (Cost Performance Index)'
        ])
        
        resumo_data['Valor'].extend([
            sum(metrics['total_cost'] for metrics in project_metrics.values()),
            sum(metrics['realized_cost'] for metrics in project_metrics.values()),
            sum(metrics['cost_percentage'] for metrics in project_metrics.values()) / len(project_metrics) if project_metrics else 0,  # Novo valor
            sum(metrics['cpi'] for metrics in project_metrics.values() if metrics['cpi'] > 0) / len(project_metrics) if project_metrics else 0
        ])
    
    # Criar DataFrame de resumo
    resumo_df = pd.DataFrame(resumo_data)
    
    # Escrever planilha de resumo
    resumo_df.to_excel(writer, sheet_name='Resumo', index=False)
    
    # Calcular distribuição por tipo de projeto
    project_type_counts = {}
    for project_id, metrics in project_metrics.items():
        project = projects_df[projects_df['project_id'] == project_id].iloc[0]
        project_type = project['project_type']
        if project_type not in project_type_counts:
            project_type_counts[project_type] = 0
        project_type_counts[project_type] += 1
    
    # Criar DataFrame com distribuição por tipo
    types_df = pd.DataFrame({
        'Tipo de Projeto': list(project_type_counts.keys()),
        'Quantidade': list(project_type_counts.values())
    }).sort_values('Quantidade', ascending=False)
    
    # Adicionar à planilha de resumo
    start_row = len(resumo_data['Métrica']) + 3
    types_df.to_excel(writer, sheet_name='Resumo', startrow=start_row, index=False)
    
    # Planilha de visão geral de projetos
    all_projects_data = []
    
    for project_id, metrics in project_metrics.items():
        # Obter dados do projeto
        project = projects_df[projects_df['project_id'] == project_id].iloc[0]
        
        # Obter nome do cliente
        client_id = project['client_id']
        client_name = clients_df[clients_df['client_id'] == client_id]['name'].iloc[0] if not clients_df.empty else "Cliente Desconhecido"
        
        # Obter nome da equipe
        team_id = project['group_id']
        team_name = groups_df[groups_df['id'] == team_id]['group_name'].iloc[0] if not groups_df.empty else "Equipe Desconhecida"
        
        # Dados do projeto
        project_data = {
            'Equipe': team_name,
            'Cliente': client_name,
            'Projeto': project['project_name'],
            'Tipo de Projeto': project['project_type'],
            'Data Início': project['start_date'],
            'Data Fim': project['end_date'],
            'Horas Atuais': metrics['realized_hours'],  # Alterado de "Realizadas" para "Atuais"
            'Horas Planejadas': metrics['total_hours'],
            'Horas Gastas (%)': metrics['hours_percentage'],  # Alterado de "Conclusão" para "Horas Gastas"
            'Dias Restantes': metrics['days_remaining'],
            'Status de Risco': metrics['risk_level'],
            'Razão do Risco': metrics['risk_reason']  # Novo campo
        }
        
        # Adicionar métricas financeiras se necessário
        if include_financial:
            project_data.update({
                'Custo Atual': metrics['realized_cost'],  # Alterado de "Realizado" para "Atual"
                'Custo Planejado': metrics['total_cost'],
                'Custo Gasto (%)': metrics['cost_percentage'],  # Novo campo
                'CPI': metrics['cpi']
            })
        
        all_projects_data.append(project_data)
    
    # Criar DataFrame com todos os projetos
    if all_projects_data:
        all_projects_df = pd.DataFrame(all_projects_data)
        all_projects_df.to_excel(writer, sheet_name='Visão Geral', index=False)
    
    # Planilhas individuais por equipe
    teams_mapping = {}
    
    # Preparar mapeamento de equipes
    if "Todas" in selected_teams:
        for _, team in groups_df.iterrows():
            teams_mapping[team['group_name']] = []
    else:
        for team in selected_teams:
            teams_mapping[team] = []
    
    # Preencher o mapeamento
    for project_id, metrics in project_metrics.items():
        project = projects_df[projects_df['project_id'] == project_id].iloc[0]
        team_id = project['group_id']
        team_name = groups_df[groups_df['id'] == team_id]['group_name'].iloc[0] if not groups_df.empty else "Equipe Desconhecida"
        
        if team_name in teams_mapping:
            teams_mapping[team_name].append({
                'project_id': project_id,
                'project': project,
                'metrics': metrics
            })
    
    # Criar planilha para cada equipe
    for team_name, team_projects in teams_mapping.items():
        if not team_projects:
            continue
            
        # Dados dos projetos da equipe
        team_data = []
        
        for project_info in team_projects:
            project = project_info['project']
            metrics = project_info['metrics']
            
            # Obter nome do cliente
            client_id = project['client_id']
            client_name = clients_df[clients_df['client_id'] == client_id]['name'].iloc[0] if not clients_df.empty else "Cliente Desconhecido"
            
            # Dados do projeto
            project_data = {
                'Cliente': client_name,
                'Projeto': project['project_name'],
                'Tipo de Projeto': project['project_type'],
                'Data Início': project['start_date'],
                'Data Fim': project['end_date'],
                'Horas Atuais': metrics['realized_hours'],  # Alterado
                'Horas Planejadas': metrics['total_hours'],
                'Horas Gastas (%)': metrics['hours_percentage'],  # Alterado
                'Dias Restantes': metrics['days_remaining'],
                'Status de Risco': metrics['risk_level'],
                'Razão do Risco': metrics['risk_reason']  # Novo campo
            }
            
            # Adicionar métricas financeiras se necessário
            if include_financial:
                project_data.update({
                    'Custo Atual': metrics['realized_cost'],  # Alterado
                    'Custo Planejado': metrics['total_cost'],
                    'Custo Gasto (%)': metrics['cost_percentage'],  # Novo campo
                    'CPI': metrics['cpi']
                })
            
            team_data.append(project_data)
        
        # Criar DataFrame para a equipe
        team_df = pd.DataFrame(team_data)
        
        # Nome seguro para a planilha (max 31 caracteres)
        sheet_name = team_name[:31] if len(team_name) > 31 else team_name
        team_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Se solicitado, adicionar planilhas com detalhes de horas por projeto
        if include_hour_details:
            for project_info in team_projects:
                project_id = project_info['project_id']
                project_name = project_info['project']['project_name']
                
                # Obter detalhes de horas por recurso
                resource_hours = get_resource_hours(
                    timesheet_df,
                    project_id,
                    users_df
                )
                
                if resource_hours:
                    # Criar DataFrame com horas por recurso
                    hours_data = [{
                        'Colaborador': resource['name'],
                        'Horas': resource['hours'],
                        'Percentual': resource['percentage']
                    } for resource in resource_hours]
                    
                    hours_df = pd.DataFrame(hours_data)
                    
                    # Nome seguro para a planilha (max 31 caracteres)
                    # Usar prefixo "Horas_" + início do nome do projeto
                    sheet_name = f"Horas_{project_name[:24]}"
                    hours_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Adicionar formatação ao arquivo Excel
    workbook = writer.book
    
    # Formatar a planilha de resumo
    worksheet = writer.sheets['Resumo']
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#1E77B4',  # Azul
        'font_color': 'white',
        'border': 1
    })
    
    metric_format = workbook.add_format({
        'border': 1
    })
    
    value_format = workbook.add_format({
        'border': 1,
        'num_format': '0.00'
    })
    
    percent_format = workbook.add_format({
        'border': 1,
        'num_format': '0.0%'
    })
    
    currency_format = workbook.add_format({
        'border': 1,
        'num_format': '€#,##0.00'
    })
    
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'font_color': '#1E77B4'
    })
    
    # Aplicar formatação à planilha de resumo
    worksheet.set_column('A:A', 30)
    worksheet.set_column('B:B', 15)
    
    # Formatar cabeçalhos
    worksheet.write(0, 0, 'Métrica', header_format)
    worksheet.write(0, 1, 'Valor', header_format)
    
    # Adicionar título para seção de tipos de projeto
    worksheet.write(start_row - 1, 0, 'Distribuição por Tipo de Projeto', title_format)
    worksheet.write(start_row, 0, 'Tipo de Projeto', header_format)
    worksheet.write(start_row, 1, 'Quantidade', header_format)
    
    # Formatar a planilha de visão geral
    if all_projects_data:
        worksheet = writer.sheets['Visão Geral']
        
        # Formatar cabeçalhos
        for col_num, value in enumerate(all_projects_df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            
        # Ajustar larguras das colunas
        worksheet.set_column('A:A', 15)  # Equipe
        worksheet.set_column('B:B', 20)  # Cliente
        worksheet.set_column('C:C', 25)  # Projeto
        worksheet.set_column('D:D', 15)  # Tipo de Projeto
        worksheet.set_column('E:F', 12)  # Datas
        worksheet.set_column('G:H', 15)  # Horas
        worksheet.set_column('I:I', 12)  # % Horas
        worksheet.set_column('J:J', 12)  # Dias Restantes
        worksheet.set_column('K:K', 15)  # Status de Risco
        worksheet.set_column('L:L', 40)  # Razão do Risco
        
        if include_financial:
            worksheet.set_column('M:N', 15)  # Custos
            worksheet.set_column('O:O', 12)  # % Custo
            worksheet.set_column('P:P', 10)  # CPI
    
    # Salvar o arquivo Excel
    writer._save()
    print(f"Excel gerado com sucesso: {output_path}")
    return True

def send_email(
    recipients, 
    subject, 
    message, 
    pdf_path, 
    excel_path, 
    smtp_server, 
    smtp_port, 
    smtp_user, 
    smtp_password, 
    use_tls
):
    """
    Envia email com os relatórios anexados
    """
    try:
        # Criar email MIME
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = ", ".join(recipients)
        msg['Subject'] = subject
        
        # Adicionar corpo do email
        msg.attach(MIMEText(message, 'plain'))
        
        # Verificar se os arquivos existem antes de anexar
        has_attachments = False
        
        # Anexar PDF se existir
        if pdf_path and os.path.exists(pdf_path):
            with open(pdf_path, "rb") as pdf_file:
                attachment = MIMEApplication(pdf_file.read(), _subtype="pdf")
                attachment.add_header(
                    'Content-Disposition', 
                    'attachment', 
                    filename="relatorio_status_projetos.pdf"
                )
                msg.attach(attachment)
                has_attachments = True
                st.success(f"PDF anexado ao email.")
        else:
            st.warning(f"Arquivo PDF não encontrado em {pdf_path}")
        
        # Anexar Excel se existir
        if excel_path and os.path.exists(excel_path):
            with open(excel_path, "rb") as excel_file:
                attachment = MIMEApplication(excel_file.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                attachment.add_header(
                    'Content-Disposition', 
                    'attachment', 
                    filename="relatorio_status_projetos.xlsx"
                )
                msg.attach(attachment)
                has_attachments = True
                st.success(f"Excel anexado ao email.")
        else:
            st.warning(f"Arquivo Excel não encontrado em {excel_path}")
            
        if not has_attachments:
            st.error("Nenhum anexo disponível para enviar. Verifique se os relatórios foram gerados corretamente.")
            return False
        
        # Configurar e enviar email
        server = smtplib.SMTP(smtp_server, smtp_port)
        
        if use_tls:
            server.starttls()
        
        if smtp_user and smtp_password:
            server.login(smtp_user, smtp_password)
        
        server.send_message(msg)
        server.quit()
        
        st.success(f"Email enviado com sucesso para {', '.join(recipients)}!")
        return True
    except Exception as e:
        import traceback
        error_msg = f"Erro ao enviar email: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())
        st.error(error_msg)
        return False

def generate_pdf_report(
    output_path, 
    projects_df, 
    timesheet_df, 
    clients_df, 
    users_df, 
    rates_df,
    groups_df,
    start_date, 
    end_date, 
    include_charts, 
    include_financial,
    include_hour_details,
    selected_teams,
    selected_clients,
    selected_project_types
):
    """
    Gera relatório PDF com status dos projetos com melhor formatação
    """
    try:
        # Criar diretório para fontes se não existir
        fonts_dir = './fonts'
        os.makedirs(fonts_dir, exist_ok=True)
        
        # Verificar se as fontes existem, caso contrário usar fontes padrão
        font_files = {
            'DejaVuSansCondensed.ttf': False,
            'DejaVuSansCondensed-Bold.ttf': False,
            'DejaVuSansCondensed-Oblique.ttf': False
        }
        
        for font_file in font_files:
            if os.path.exists(os.path.join(fonts_dir, font_file)):
                font_files[font_file] = True
        
        # Se pelo menos uma fonte não existir, usar PDF padrão
        use_improved_pdf = all(font_files.values())
        
        if use_improved_pdf:
            pdf = ImprovedPDF()
        else:
            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=20)  # Aumentei a margem para evitar cortes
        
        # Calcular métricas para cada projeto
        project_metrics = calculate_project_metrics(
            projects_df, 
            timesheet_df, 
            clients_df, 
            users_df, 
            rates_df
        )
        
        # Primeira página - Capa
        pdf.add_page()
        
        # Cabeçalho da capa
        if os.path.exists('logo.png'):
            pdf.image('logo.png', x=20, y=20, w=40)
        
        pdf.set_font('Arial', 'B', 22)
        pdf.set_text_color(31, 119, 180)  # Azul
        pdf.set_xy(20, 60)
        pdf.cell(170, 15, 'Relatório Executivo de Projetos', 0, 1, 'C')
        
        pdf.set_font('Arial', 'B', 16)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(20, 80)
        pdf.cell(170, 10, f"Status e Métricas de Desempenho", 0, 1, 'C')
        
        # Informações do período
        pdf.set_font('Arial', '', 12)
        pdf.set_xy(20, 100)
        pdf.cell(170, 10, f"Período: {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}", 0, 1, 'C')
        
        # Linha divisória
        pdf.set_draw_color(31, 119, 180)  # Azul
        pdf.set_line_width(0.5)
        pdf.line(50, 115, 160, 115)
        
        # Filtros aplicados
        pdf.set_font('Arial', 'B', 12)
        pdf.set_xy(20, 125)
        pdf.cell(170, 10, 'Filtros Aplicados', 0, 1, 'C')
        
        pdf.set_font('Arial', '', 11)
        
        # Texto de equipas
        teams_str = ', '.join(selected_teams) if "Todas" not in selected_teams else "Todas as equipes"
        pdf.set_xy(50, 140)
        pdf.cell(110, 8, f"Equipes: {teams_str}", 0, 1, 'L')
        
        # Texto de clientes
        clients_str = ', '.join(selected_clients) if "Todos" not in selected_clients else "Todos os clientes"
        pdf.set_xy(50, 150)
        pdf.cell(110, 8, f"Clientes: {clients_str}", 0, 1, 'L')
        
        # Texto de tipos de projeto (novo)
        types_str = ', '.join(selected_project_types) if "Todos" not in selected_project_types else "Todos os tipos"
        pdf.set_xy(50, 160)
        pdf.cell(110, 8, f"Tipos de Projeto: {types_str}", 0, 1, 'L')
        
        # Data de geração
        pdf.set_font('Arial', 'I', 10)
        pdf.set_text_color(100, 100, 100)  # Cinza escuro
        pdf.set_xy(20, 250)
        pdf.cell(170, 10, f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", 0, 1, 'C')
        
        # Organizar projetos por equipe
        teams_mapping = {}
        
        # Se selecionou "Todas", incluir todas as equipes
        if "Todas" in selected_teams:
            for _, team in groups_df.iterrows():
                teams_mapping[team['group_name']] = []
        else:
            for team in selected_teams:
                teams_mapping[team] = []
        
        # Preencher o mapeamento com os projetos de cada equipe
        for _, project in projects_df.iterrows():
            # Obter nome da equipe - com tratamento de erro mais robusto
            team_id = project['group_id']
            team_name = "Equipe Desconhecida"  # Valor padrão

            # Verificar se o DataFrame de grupos não está vazio
            if not groups_df.empty:
                # Filtrar para encontrar o grupo correspondente
                filtered_groups = groups_df[groups_df['id'] == team_id]
                
                # Verificar se a filtragem retornou algum resultado
                if not filtered_groups.empty:
                    team_name = filtered_groups['group_name'].iloc[0]
            
            if team_name in teams_mapping:
                # Verificar se o projeto está nas métricas calculadas
                project_id = project['project_id']
                if project_id in project_metrics:
                    teams_mapping[team_name].append({
                        'project': project,
                        'metrics': project_metrics[project_id]
                    })
        
        # Resumo Geral
        pdf.add_page()
        
        if use_improved_pdf:
            pdf.chapter_title("Resumo Geral")
        else:
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(190, 10, "Resumo Geral", 0, 1, 'L')
        
        # Tabela de resumo
        pdf.set_fill_color(240, 240, 240)
        pdf.set_font('Arial', 'B', 10)
        
        # Cabeçalho da tabela
        pdf.set_fill_color(31, 119, 180)  # Azul
        pdf.set_text_color(255, 255, 255)  # Branco
        pdf.cell(70, 10, "Métrica", 1, 0, 'L', True)
        pdf.cell(30, 10, "Valor", 1, 1, 'C', True)
        
        # Total de projetos
        pdf.set_font('Arial', '', 10)
        pdf.set_text_color(0, 0, 0)  # Preto
        pdf.set_fill_color(255, 255, 255)  # Fundo branco
        pdf.cell(70, 8, "Total de Projetos Ativos", 1, 0, 'L')
        pdf.cell(30, 8, str(len(project_metrics)), 1, 1, 'C')
        
        # Total de horas registradas
        pdf.set_fill_color(245, 245, 245)  # Cinza claro para alternar linhas
        total_hours = sum(metrics['realized_hours'] for metrics in project_metrics.values())
        pdf.cell(70, 8, "Total de Horas Atuais", 1, 0, 'L', True)  # Alterado de "Registradas" para "Atuais"
        pdf.cell(30, 8, f"{total_hours:.2f}h", 1, 1, 'C', True)
        
        # Total de horas planejadas
        total_planned_hours = sum(metrics['total_hours'] for metrics in project_metrics.values())
        pdf.cell(70, 8, "Total de Horas Planejadas", 1, 0, 'L')
        pdf.cell(30, 8, f"{total_planned_hours:.2f}h", 1, 1, 'C')
        
        # Percentual médio de horas gastas (em vez de conclusão)
        pdf.set_fill_color(245, 245, 245)  # Cinza claro para alternar linhas
        avg_hours_percent = sum(metrics['hours_percentage'] for metrics in project_metrics.values()) / len(project_metrics) if project_metrics else 0
        pdf.cell(70, 8, "Percentual Médio de Horas Gastas", 1, 0, 'L', True)  # Alterado
        pdf.cell(30, 8, f"{avg_hours_percent:.1f}%", 1, 1, 'C', True)
        
        if include_financial:
            # Custo total planejado
            total_planned_cost = sum(metrics['total_cost'] for metrics in project_metrics.values())
            pdf.cell(70, 8, "Custo Total Planejado", 1, 0, 'L')
            pdf.cell(30, 8, f"EUR {total_planned_cost:.2f}", 1, 1, 'C')
            
            # Custo total realizado
            pdf.set_fill_color(245, 245, 245)  # Cinza claro para alternar linhas
            total_realized_cost = sum(metrics['realized_cost'] for metrics in project_metrics.values())
            pdf.cell(70, 8, "Custo Total Atual", 1, 0, 'L', True)  # Alterado de "Realizado" para "Atual"
            pdf.cell(30, 8, f"EUR {total_realized_cost:.2f}", 1, 1, 'C', True)
            
            # Percentual médio de custo gasto
            avg_cost_percent = sum(metrics['cost_percentage'] for metrics in project_metrics.values()) / len(project_metrics) if project_metrics else 0
            pdf.cell(70, 8, "Percentual Médio de Custo Gasto", 1, 0, 'L')  # Novo campo
            pdf.cell(30, 8, f"{avg_cost_percent:.1f}%", 1, 1, 'C')
            
            # CPI médio
            pdf.set_fill_color(245, 245, 245)  # Cinza claro para alternar linhas
            avg_cpi = sum(metrics['cpi'] for metrics in project_metrics.values() if metrics['cpi'] > 0) / len(project_metrics) if project_metrics else 0
            pdf.cell(70, 8, "CPI Médio (Cost Performance Index)", 1, 0, 'L', True)
            pdf.cell(30, 8, f"{avg_cpi:.2f}", 1, 1, 'C', True)
        
        # Distribuição por tipo de projeto (novo)
        # Contar projetos por tipo
        project_type_counts = {}
        for project_id, metrics in project_metrics.items():
            project = projects_df[projects_df['project_id'] == project_id].iloc[0]
            project_type = project['project_type']
            if project_type not in project_type_counts:
                project_type_counts[project_type] = 0
            project_type_counts[project_type] += 1
        
        # Verificar se há espaço suficiente para a tabela de distribuição por tipo
        # Se não houver, adicionar uma nova página
        if pdf.get_y() > 220:  # Se estiver perto do fim da página
            pdf.add_page()
        else:
            pdf.ln(10)  # Espaço extra se continuar na mesma página
        
        # Tabela de distribuição por tipo
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(190, 10, "Distribuição por Tipo de Projeto", 0, 1, 'L')
        
        pdf.set_font('Arial', 'B', 10)
        pdf.set_fill_color(31, 119, 180)  # Azul
        pdf.set_text_color(255, 255, 255)  # Branco
        pdf.cell(100, 10, "Tipo de Projeto", 1, 0, 'L', True)
        pdf.cell(30, 10, "Quantidade", 1, 1, 'C', True)
        
        pdf.set_font('Arial', '', 10)
        pdf.set_text_color(0, 0, 0)  # Preto
        
        alternating = False
        for project_type, count in sorted(project_type_counts.items(), key=lambda x: x[1], reverse=True):
            if alternating:
                pdf.set_fill_color(245, 245, 245)  # Cinza claro
            else:
                pdf.set_fill_color(255, 255, 255)  # Branco
            
            pdf.cell(100, 8, project_type, 1, 0, 'L', alternating)
            pdf.cell(30, 8, str(count), 1, 1, 'C', alternating)
            alternating = not alternating
        
        # Distribuição de projetos por status
        if include_charts:
            # Criar gráficos no diretório temporário
            temp_dir = os.path.dirname(output_path)
            
            # Classificar projetos por status de risco
            risk_counts = {'Alto': 0, 'Médio': 0, 'Baixo': 0}
            for metrics in project_metrics.values():
                # Ajustar chaves para evitar acentos
                risk_level = metrics['risk_level']
                if risk_level == 'Médio':
                    risk_level = 'Médio'
                risk_counts[risk_level] += 1
            
            # Criar dataframe para o gráfico de pizza
            risk_df = pd.DataFrame({
                'Status': list(risk_counts.keys()),
                'Quantidade': list(risk_counts.values())
            })
            
            # Criar gráfico de pizza de distribuição de risco
            if sum(risk_counts.values()) > 0:
                fig_risk = px.pie(
                    risk_df, 
                    values='Quantidade', 
                    names='Status', 
                    title='Distribuição de Projetos por Status de Risco',
                    color='Status',
                    color_discrete_map={'Alto': '#FF5252', 'Médio': '#FFC107', 'Baixo': '#4CAF50'},
                    hole=0.4  # Adicionar um buraco no meio para um visual moderno
                )
                
                # Melhorar layout
                fig_risk.update_layout(
                    legend_title_text='Status de Risco',
                    font=dict(size=12),
                    title_font=dict(size=16, color='#444'),
                    title_x=0.5,  # Centralizar o título
                    margin=dict(t=50, b=20, l=20, r=20)
                )
                
                risk_chart_path = os.path.join(temp_dir, "risk_distribution.png")
                fig_risk.write_image(risk_chart_path, width=600, height=400)
                
                # Adicionar nova página para o gráfico
                pdf.add_page()
                pdf.set_font('Arial', 'B', 14)
                pdf.cell(190, 10, "Distribuição de Status de Risco", 0, 1, 'L')
                pdf.image(risk_chart_path, x=20, y=40, w=170)
                
                # Legenda
                pdf.set_y(150)
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(190, 10, "Legenda de Status de Risco:", 0, 1)
                
                pdf.set_font('Arial', '', 10)
                
                # Vermelho - Alto Risco
                pdf.set_fill_color(255, 82, 82)  # #FF5252
                pdf.rect(20, 160, 5, 5, 'F')
                pdf.set_xy(30, 160)
                pdf.cell(100, 5, f"Alto Risco: {risk_counts['Alto']} projetos", 0, 1)
                
                # Amarelo - Médio Risco
                pdf.set_fill_color(255, 193, 7)  # #FFC107
                pdf.rect(20, 170, 5, 5, 'F')
                pdf.set_xy(30, 170)
                pdf.cell(100, 5, f"Médio Risco: {risk_counts['Médio']} projetos", 0, 1)
                
                # Verde - Baixo Risco
                pdf.set_fill_color(76, 175, 80)  # #4CAF50
                pdf.rect(20, 180, 5, 5, 'F')
                pdf.set_xy(30, 180)
                pdf.cell(100, 5, f"Baixo Risco: {risk_counts['Baixo']} projetos", 0, 1)
                
                pdf.ln(10)
                
                # Explicação sobre risco
                pdf.set_font('Arial', 'I', 9)
                pdf.set_xy(20, 190)
                pdf.multi_cell(
                    170, 
                    5, 
                    "Nota: O nível de risco é calculado com base no CPI (Cost Performance Index) e no progresso do projeto em relação ao cronograma. Um CPI maior ou igual a 1.1 indica Baixo Risco, entre 0.9 e 1.1 indica Médio Risco, e menor que 0.9 indica Alto Risco."
                )
            
            # Criar gráfico de barras para tipos de projeto
            if project_type_counts:
                types_df = pd.DataFrame({
                    'Tipo': list(project_type_counts.keys()),
                    'Quantidade': list(project_type_counts.values())
                })
                
                fig_types = px.bar(
                    types_df.sort_values('Quantidade', ascending=False), 
                    x='Tipo', 
                    y='Quantidade',
                    title='Distribuição de Projetos por Tipo',
                    color='Quantidade',
                    text='Quantidade',
                    color_continuous_scale='Viridis'
                )
                
                # Melhorar layout
                fig_types.update_layout(
                    xaxis_title='Tipo de Projeto',
                    yaxis_title='Quantidade de Projetos',
                    font=dict(size=12),
                    title_font=dict(size=16, color='#444'),
                    title_x=0.5,
                    margin=dict(t=50, b=50, l=50, r=20)
                )
                
                types_chart_path = os.path.join(temp_dir, "project_types.png")
                fig_types.write_image(types_chart_path, width=600, height=400)
                
                # Adicionar gráfico ao PDF em nova página
                pdf.add_page()
                pdf.set_font('Arial', 'B', 14)
                pdf.cell(190, 10, "Distribuição por Tipo de Projeto", 0, 1, 'L')
                pdf.image(types_chart_path, x=20, y=40, w=170)
        
        # Para cada equipe, listar os projetos com suas métricas
        for team_name, team_projects in teams_mapping.items():
            if team_projects:
                pdf.add_page()
                
                if use_improved_pdf:
                    pdf.chapter_title(f"Equipe: {team_name}")
                else:
                    pdf.set_font('Arial', 'B', 14)
                    pdf.set_fill_color(220, 220, 220)  # Cinza claro para o fundo
                    pdf.cell(190, 10, f"Equipe: {team_name}", 0, 1, 'L', True)
                
                # Listar projetos da equipe
                for project_data in team_projects:
                    project = project_data['project']
                    metrics = project_data['metrics']
                    
                    # Verificar se este projeto começaria muito próximo ao fim da página
                    if pdf.get_y() > 220:  # Se estiver perto do fim da página
                        pdf.add_page()
                    
                    # Nome do cliente
                    client_id = project['client_id']
                    client_name = clients_df[clients_df['client_id'] == client_id]['name'].iloc[0] if not clients_df.empty else "Cliente Desconhecido"
                    
                    # Cabeçalho do projeto - com área sombreada
                    pdf.ln(5)
                    pdf.set_fill_color(240, 240, 240)  # Cinza muito claro
                    pdf.set_font('Arial', 'B', 12)
                    pdf.cell(190, 10, f"Projeto: {project['project_name']}", 0, 1, 'L', True)
                    
                    # Informações básicas
                    pdf.set_font('Arial', '', 10)
                    pdf.cell(40, 7, "Cliente:", 0, 0)
                    pdf.set_font('Arial', 'B', 10)
                    pdf.cell(150, 7, client_name, 0, 1)
                    
                    pdf.set_font('Arial', '', 10)
                    pdf.cell(40, 7, "Tipo de Projeto:", 0, 0)
                    pdf.set_font('Arial', 'B', 10)
                    pdf.cell(150, 7, project['project_type'], 0, 1)
                    
                    pdf.set_font('Arial', '', 10)
                    pdf.cell(40, 7, "Período:", 0, 0)
                    pdf.set_font('Arial', '', 10)
                    pdf.cell(150, 7, f"{project['start_date'].strftime('%d/%m/%Y')} a {project['end_date'].strftime('%d/%m/%Y')}", 0, 1)
                    
                    # Status geral do projeto com indicador de cor
                    status_color = {'Alto': (255, 82, 82), 'Médio': (255, 193, 7), 'Baixo': (76, 175, 80)}
                    
                    # Fundo colorido para o status
                    risk_key = metrics['risk_level']
                    if risk_key not in status_color:
                        if risk_key == 'Médio':
                            risk_key = 'Médio'
                    
                    pdf.set_font('Arial', '', 10)
                    pdf.cell(40, 7, "Status de Risco:", 0, 0)
                    
                    # Retângulo colorido para status
                    risk_color = status_color[risk_key]
                    x_pos = pdf.get_x()
                    y_pos = pdf.get_y()
                    pdf.set_fill_color(*risk_color)
                    pdf.rect(x_pos, y_pos, 30, 7, 'F')
                    
                    # Texto branco sobre o retângulo colorido
                    pdf.set_text_color(255, 255, 255)
                    pdf.set_xy(x_pos, y_pos)
                    pdf.cell(30, 7, risk_key, 0, 1, 'C')
                    pdf.set_text_color(0, 0, 0)  # Restaurar cor
                    
                    # Razão do status de risco (novo)
                    pdf.set_font('Arial', '', 10)
                    pdf.cell(40, 7, "Razão do Risco:", 0, 0)
                    pdf.set_font('Arial', 'I', 9)
                    
                    # Verificar espaço disponível para a razão do risco
                    risk_reason = metrics['risk_reason']
                    if len(risk_reason) > 70:  # Se for muito longo
                        # Dividir em duas linhas
                        x_pos = pdf.get_x()
                        y_pos = pdf.get_y()
                        pdf.multi_cell(150, 7, risk_reason, 0, 'L')
                    else:
                        pdf.cell(150, 7, risk_reason, 0, 1)
                    
                    # Métricas do projeto - Tabela
                    pdf.ln(3)
                    
                    # Verificar se a tabela excede o limite da página
                    if pdf.get_y() > 220:  # Se estiver perto do fim da página
                        pdf.add_page()
                    
                    # Cabeçalho da tabela
                    pdf.set_fill_color(31, 119, 180)  # Azul
                    pdf.set_text_color(255, 255, 255)  # Branco
                    pdf.set_font('Arial', 'B', 10)
                    pdf.cell(90, 8, "Métrica", 1, 0, 'L', True)
                    pdf.cell(45, 8, "Valor Atual", 1, 0, 'C', True)
                    pdf.cell(45, 8, "Total Planejado", 1, 1, 'C', True)
                    
                    # Restaurar cores de texto
                    pdf.set_text_color(0, 0, 0)
                    
                    # Horas - agora com valor atual e total planejado
                    pdf.set_font('Arial', '', 10)
                    pdf.cell(90, 7, "Horas", 1, 0, 'L')
                    pdf.cell(45, 7, f"{metrics['realized_hours']:.2f}h", 1, 0, 'C')
                    pdf.cell(45, 7, f"{metrics['total_hours']:.2f}h", 1, 1, 'C')
                    
                    # Percentual de horas gastas (em vez de conclusão)
                    pdf.set_fill_color(245, 245, 245)  # Cinza claro
                    pdf.cell(90, 7, "Percentual de Horas Gastas", 1, 0, 'L', True)  # Alterado
                    pdf.cell(45, 7, f"{metrics['hours_percentage']:.1f}%", 1, 0, 'C', True)
                    pdf.cell(45, 7, "100%", 1, 1, 'C', True)
                    
                    # Dias restantes
                    pdf.cell(90, 7, "Dias Restantes", 1, 0, 'L')
                    pdf.cell(45, 7, f"{metrics['days_remaining']}", 1, 0, 'C')
                    pdf.cell(45, 7, f"{metrics['total_days']}", 1, 1, 'C')
                    
                    if include_financial:
                        # Custo - agora com valor atual e total planejado
                        pdf.set_fill_color(245, 245, 245)  # Cinza claro
                        pdf.cell(90, 7, "Custo", 1, 0, 'L', True)
                        pdf.cell(45, 7, f"EUR {metrics['realized_cost']:.2f}", 1, 0, 'C', True)
                        pdf.cell(45, 7, f"EUR {metrics['total_cost']:.2f}", 1, 1, 'C', True)
                        
                        # Percentual de custo gasto (novo)
                        pdf.cell(90, 7, "Percentual de Custo Gasto", 1, 0, 'L')  # Novo campo
                        pdf.cell(45, 7, f"{metrics['cost_percentage']:.1f}%", 1, 0, 'C')
                        pdf.cell(45, 7, "100%", 1, 1, 'C')
                        
                        # CPI
                        pdf.set_fill_color(245, 245, 245)  # Cinza claro
                        pdf.cell(90, 7, "CPI (Cost Performance Index)", 1, 0, 'L', True)
                        
                        # Definir cor do CPI
                        if metrics['cpi'] >= 1.1:
                            pdf.set_text_color(0, 128, 0)  # Verde
                        elif metrics['cpi'] >= 0.9:
                            pdf.set_text_color(255, 128, 0)  # Laranja
                        else:
                            pdf.set_text_color(255, 0, 0)  # Vermelho
                        
                        pdf.cell(45, 7, f"{metrics['cpi']:.2f}", 1, 0, 'C', True)
                        pdf.set_text_color(0, 0, 0)  # Restaurar cor
                        pdf.cell(45, 7, ">= 1.0", 1, 1, 'C', True)
                    
                    # Adicionar detalhamento de horas por recurso se solicitado
                    if include_hour_details:
                        # Obtém detalhes das horas por recurso
                        resource_hours = get_resource_hours(
                            timesheet_df, 
                            project['project_id'], 
                            users_df
                        )
                        
                        if resource_hours:
                            # Verificar se há espaço suficiente na página atual
                            if pdf.get_y() > 220:  # Se estiver próximo do fim da página
                                pdf.add_page()
                                pdf.set_font('Arial', 'B', 11)
                                pdf.cell(190, 10, f"Detalhamento de Horas - {project['project_name']}", 0, 1, 'L')
                            else:
                                pdf.ln(5)
                                pdf.set_font('Arial', 'B', 11)
                                pdf.set_fill_color(220, 220, 220)  # Cinza claro
                                pdf.cell(190, 8, "Detalhamento de Horas por Recurso", 0, 1, 'L', True)
                            
                            # Cabeçalho da tabela de horas
                            pdf.set_fill_color(31, 119, 180)  # Azul
                            pdf.set_text_color(255, 255, 255)  # Branco
                            pdf.set_font('Arial', 'B', 9)
                            pdf.cell(80, 7, "Colaborador", 1, 0, 'L', True)
                            pdf.cell(35, 7, "Horas", 1, 0, 'C', True)
                            pdf.cell(35, 7, "% do Total", 1, 1, 'C', True)
                            
                            # Dados da tabela
                            pdf.set_text_color(0, 0, 0)  # Preto
                            pdf.set_font('Arial', '', 9)
                            
                            alternating = False
                            for resource in resource_hours:
                                if alternating:
                                    pdf.set_fill_color(245, 245, 245)  # Cinza claro
                                else:
                                    pdf.set_fill_color(255, 255, 255)  # Branco
                                
                                pdf.cell(80, 6, resource['name'], 1, 0, 'L', alternating)
                                pdf.cell(35, 6, f"{resource['hours']:.2f}h", 1, 0, 'C', alternating)
                                pdf.cell(35, 6, f"{resource['percentage']:.1f}%", 1, 1, 'C', alternating)
                                alternating = not alternating
        
        # Salvar o PDF
        pdf.output(output_path)
        st.success(f"PDF gerado com sucesso: {output_path}")
        return True
    except Exception as e:
        import traceback
        error_msg = f"Erro ao gerar PDF: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())
        st.error(error_msg)
        return False